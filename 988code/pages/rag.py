from .common import *
from dash import callback_context
from dash.exceptions import PreventUpdate
from dash import ALL, no_update
import datetime
import openpyxl
import xlrd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import tempfile
import base64
import os

from env_loader import get_env_int

# TODO 現在還不知道檔案要存到哪

# 從資料庫載入初始條目
def load_initial_items():
    try:
        import requests
        response = requests.get("http://127.0.0.1:8000/get_rag_titles")
        if response.status_code == 200:
            data = response.json()
            titles = [item['title'] for item in data]
            return titles
        else:
            print(f"[ERROR] 載入RAG條目失敗: {response.status_code}")
            return []
    except Exception as e:
        print(f"[ERROR] 載入RAG條目失敗: {e}")
        return []

# 不在模組載入時執行，改為在 layout 中動態載入

# 儲存要被刪除的條目名稱
item_to_delete = None

def setup_chinese_font():
    """設定中文字體支援"""
    try:
        # 嘗試註冊系統中文字體
        # Windows 系統字體路徑
        font_paths = [
            r"C:\Windows\Fonts\msjh.ttc",  # 微軟正黑體
            r"C:\Windows\Fonts\SimHei.ttf",  # 黑體
            r"C:\Windows\Fonts\simsun.ttc",  # 新細明體
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    if font_path.endswith('.ttc'):
                        # TTC 字體需要指定子字體索引
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path, subfontIndex=0))
                    else:
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                    print(f"成功載入中文字體: {font_path}")
                    return True
                except Exception as e:
                    print(f"載入字體 {font_path} 失敗: {e}")
                    continue
        
        print("警告: 找不到中文字體，將使用預設字體")
        return False
    except Exception as e:
        print(f"設定中文字體時發生錯誤: {e}")
        return False

def excel_to_pdf_bytes_win32com(excel_content, filename=None):
    """
    使用 win32com 將 Excel 內容轉換為 PDF 位元組 (完美保留格式)
    
    參數:
    excel_content: Excel 檔案的位元組內容
    filename: 原始檔案名稱
    
    返回:
    PDF 檔案的位元組內容，轉換失敗則返回 None
    """
    
    try:
        import win32com.client
        import tempfile
        import os
        import datetime
        
        # 判斷檔案類型
        is_old_format = filename and filename.lower().endswith('.xls')
        suffix = '.xls' if is_old_format else '.xlsx'
        
        # 建立臨時 Excel 檔案
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp_excel:
            temp_excel_path = temp_excel.name
            temp_excel.write(excel_content)
            temp_excel.flush()
            
        # 建立臨時 PDF 檔案路徑
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
            temp_pdf_path = temp_pdf.name
            
        print(f"使用 win32com 轉換: {filename}")
        print(f"臨時 Excel: {temp_excel_path}")
        print(f"臨時 PDF: {temp_pdf_path}")
        
        try:
            # 創建 Excel 應用程序對象
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # 不顯示 Excel 窗口
            excel.DisplayAlerts = False  # 不顯示警告對話框
            
            # 打開 Excel 文件
            workbook = excel.Workbooks.Open(os.path.abspath(temp_excel_path))
            
            # 為確保每次轉換都不同，在第一個工作表加入隱藏註解
            try:
                worksheet = workbook.Worksheets(1)
                current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
                # 在最後一個空白儲存格加入時間戳註解（白色字體，不可見）
                last_row = worksheet.UsedRange.Rows.Count + 1
                cell = worksheet.Cells(last_row, 1)
                cell.Value = f"<!-- {current_time} -->"
                cell.Font.Color = 16777215  # 白色字體（不可見）
                cell.Font.Size = 1  # 極小字體
            except:
                pass  # 如果加入註解失敗，繼續進行轉換
            
            # 導出為 PDF (0 = xlTypePDF)
            workbook.ExportAsFixedFormat(
                Type=0,  # xlTypePDF
                Filename=os.path.abspath(temp_pdf_path),
                Quality=0,  # xlQualityStandard
                IncludeDocProps=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            # 關閉文件和應用程序
            workbook.Close(SaveChanges=False)
            excel.Quit()
            
            # 讀取生成的 PDF
            with open(temp_pdf_path, 'rb') as pdf_file:
                pdf_bytes = pdf_file.read()
            
            print(f"win32com 轉換成功，PDF 大小: {len(pdf_bytes)} bytes")
            
            # 清理臨時檔案
            try:
                os.unlink(temp_excel_path)
                os.unlink(temp_pdf_path)
            except:
                pass
            
            return pdf_bytes
            
        except Exception as excel_error:
            print(f"Excel 應用程式轉換失敗: {excel_error}")
            # 確保清理 Excel 程序
            try:
                excel.Quit()
            except:
                pass
            
            # 清理臨時檔案
            try:
                os.unlink(temp_excel_path)
                os.unlink(temp_pdf_path)
            except:
                pass
            
            return None
            
    except ImportError:
        print("win32com 模組未安裝，請執行: pip install pywin32")
        return None
    except Exception as e:
        print(f"win32com 轉換過程中發生錯誤: {str(e)}")
        return None

def excel_to_pdf_bytes(excel_content, filename=None, sheet_name=None):
    """
    將 Excel 內容轉換為 PDF 位元組
    
    參數:
    excel_content: Excel 檔案的位元組內容
    filename: 原始檔案名稱（用來判斷是 .xls 還是 .xlsx）
    sheet_name: 指定工作表名稱，如果為 None 則使用第一個工作表
    
    返回:
    PDF 檔案的位元組內容，轉換失敗則返回 None
    """
    
    try:
        # 設定中文字體
        has_chinese_font = setup_chinese_font()
        
        # 判斷檔案類型
        is_old_format = filename and filename.lower().endswith('.xls')
        
        all_sheets_data = {}
        
        if is_old_format:
            # 處理 .xls 檔案
            temp_excel_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as temp_excel:
                    temp_excel_path = temp_excel.name
                    temp_excel.write(excel_content)
                    temp_excel.flush()
                
                # 使用 xlrd 讀取 .xls 檔案
                workbook = xlrd.open_workbook(temp_excel_path)
                
                # 處理所有工作表
                sheet_names = workbook.sheet_names()
                print(f"發現 {len(sheet_names)} 個工作表: {', '.join(sheet_names)}")
                
                for sheet_idx, name in enumerate(sheet_names):
                    worksheet = workbook.sheet_by_index(sheet_idx)
                    print(f"正在處理工作表 {sheet_idx + 1}/{len(sheet_names)}: {name}")
                    
                    # 讀取工作表資料
                    sheet_data = []
                    max_row = worksheet.nrows
                    max_col = worksheet.ncols
                    
                    if max_row > 0 and max_col > 0:
                        print(f"  資料範圍: {max_row} 行 x {max_col} 列")
                        
                        # 讀取資料和格式資訊
                        for row in range(max_row):
                            row_data = []
                            for col in range(max_col):
                                try:
                                    cell_value = worksheet.cell_value(row, col)
                                    if cell_value is None or cell_value == "":
                                        cell_value = ""
                                    # 處理日期格式
                                    if worksheet.cell_type(row, col) == xlrd.XL_CELL_DATE:
                                        date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                                        cell_value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                                    row_data.append(str(cell_value))
                                except:
                                    row_data.append("")
                            sheet_data.append(row_data)
                        
                        # 對於 .xls 檔案，格式資訊較少，使用基本設定
                        all_sheets_data[name] = {
                            'data': sheet_data,
                            'col_widths': None,  # .xls 格式較難獲取精確的欄寬
                            'formatting': None   # .xls 格式的格式資訊有限
                        }
                    else:
                        print(f"  工作表 {name} 無資料，跳過")
                
                # xlrd workbook 不需要顯式關閉
                
            finally:
                # 安全地刪除臨時檔案
                if temp_excel_path and os.path.exists(temp_excel_path):
                    try:
                        os.unlink(temp_excel_path)
                    except OSError as e:
                        print(f"警告: 無法刪除臨時檔案 {temp_excel_path}: {e}")
                
        else:
            # 處理 .xlsx 檔案
            temp_excel_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
                    temp_excel_path = temp_excel.name
                    temp_excel.write(excel_content)
                    temp_excel.flush()
                
                # 讀取 Excel 檔案
                workbook = openpyxl.load_workbook(temp_excel_path)
                
                # 處理所有工作表
                sheet_names = workbook.sheetnames
                print(f"發現 {len(sheet_names)} 個工作表: {', '.join(sheet_names)}")
                
                for sheet_idx, name in enumerate(sheet_names):
                    worksheet = workbook[name]
                    print(f"正在處理工作表 {sheet_idx + 1}/{len(sheet_names)}: {name}")
                    
                    # 讀取工作表資料
                    max_col = worksheet.max_column
                    max_row = worksheet.max_row
                    
                    if max_row > 0 and max_col > 0:
                        print(f"  資料範圍: {max_row} 行 x {max_col} 列")
                        
                        sheet_data = []
                        cell_formats = []  # 儲存每個儲存格的格式資訊
                        col_widths = []    # 儲存欄寬資訊
                        
                        # 讀取欄寬資訊
                        for col in range(1, max_col + 1):
                            col_letter = worksheet.cell(row=1, column=col).column_letter
                            col_dimension = worksheet.column_dimensions.get(col_letter)
                            if col_dimension and col_dimension.width:
                                # Excel 欄寬單位轉換為點數 (大約)
                                width_points = col_dimension.width * 7.5
                            else:
                                width_points = 60  # 預設寬度
                            col_widths.append(width_points)
                        
                        # 讀取資料和格式資訊
                        for row in range(1, max_row + 1):
                            row_data = []
                            row_formats = []
                            for col in range(1, max_col + 1):
                                cell = worksheet.cell(row=row, column=col)
                                cell_value = cell.value if cell.value is not None else ""
                                row_data.append(str(cell_value))
                                
                                # 讀取儲存格格式
                                cell_format = {
                                    'font_size': cell.font.size if cell.font.size else 11,
                                    'bold': cell.font.bold if cell.font.bold else False,
                                    'italic': cell.font.italic if cell.font.italic else False,
                                    'font_name': cell.font.name if cell.font.name else 'Arial',
                                    'alignment': cell.alignment.horizontal if cell.alignment.horizontal else 'general',
                                    'fill_color': None,
                                    'font_color': None
                                }
                                
                                # 讀取背景顏色
                                if cell.fill.fill_type == 'solid' and hasattr(cell.fill, 'start_color'):
                                    if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                                        rgb = cell.fill.start_color.rgb
                                        if rgb and rgb != '00000000' and rgb != 'FFFFFFFF':  # 排除預設色
                                            cell_format['fill_color'] = rgb
                                
                                # 讀取字體顏色
                                if cell.font.color and hasattr(cell.font.color, 'rgb'):
                                    if cell.font.color.rgb and cell.font.color.rgb != '00000000':
                                        cell_format['font_color'] = cell.font.color.rgb
                                
                                row_formats.append(cell_format)
                            
                            sheet_data.append(row_data)
                            cell_formats.append(row_formats)
                        
                        all_sheets_data[name] = {
                            'data': sheet_data,
                            'col_widths': col_widths,
                            'formatting': cell_formats
                        }
                    else:
                        print(f"  工作表 {name} 無資料，跳過")
                
                # 確保 workbook 被正確關閉
                workbook.close()
                
            finally:
                # 安全地刪除臨時檔案
                if temp_excel_path and os.path.exists(temp_excel_path):
                    try:
                        os.unlink(temp_excel_path)
                    except OSError as e:
                        print(f"警告: 無法刪除臨時檔案 {temp_excel_path}: {e}")
        
        # 檢查是否有資料
        if not all_sheets_data:
            print("警告: Excel 檔案中沒有資料")
            return None
            
        # 創建臨時 PDF 檔案
        temp_pdf_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
                temp_pdf_path = temp_pdf.name
                print(f"建立臨時 PDF 檔案: {temp_pdf_path}")
            
            # 建立 PDF 文件
            doc = SimpleDocTemplate(temp_pdf_path, pagesize=A4)
            story = []
            
            # 獲取樣式
            styles = getSampleStyleSheet()
            
            # 在 PDF 中加入隱藏的時間戳註解確保唯一性
            import datetime
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
            # 加入一個不可見的元素包含時間戳
            from reportlab.platypus import Paragraph
            hidden_timestamp = Paragraph(f"<!-- {current_time} -->", styles['Normal'])
            hidden_timestamp.style.fontSize = 0.1  # 極小字體
            hidden_timestamp.style.textColor = colors.white  # 白色文字（不可見）
            story.append(hidden_timestamp)
            
            # 時間戳已透過隱藏元素加入，確保每次轉換的唯一性
            
            # 處理每個工作表
            for sheet_idx, (sheet_name, sheet_info) in enumerate(all_sheets_data.items()):
                print(f"生成 PDF 工作表 {sheet_idx + 1}/{len(all_sheets_data)}: {sheet_name}")
                
                # 不添加工作表標題，直接處理表格內容
                if isinstance(sheet_info, dict):
                    data = sheet_info['data']
                    col_widths = sheet_info['col_widths']
                    formatting = sheet_info['formatting']
                else:
                    # 向後兼容舊格式
                    data = sheet_info
                    col_widths = None
                    formatting = None
                
                # 建立表格，套用欄寬
                if col_widths:
                    table = Table(data, colWidths=col_widths)
                else:
                    table = Table(data)
                
                # 動態設定表格樣式，套用 Excel 格式
                style_commands = []
                
                # 基本邊框和對齊
                style_commands.extend([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ])
                
                def hex_to_color(hex_color):
                    """將 Excel 的 hex 顏色轉換為 ReportLab 顏色"""
                    try:
                        if hex_color and len(hex_color) >= 6:
                            # 移除 FF 前綴（如果有的話）
                            if len(hex_color) == 8 and hex_color.startswith('FF'):
                                hex_color = hex_color[2:]
                            elif len(hex_color) == 8:
                                hex_color = hex_color[2:]  # 移除 alpha 通道
                            
                            r = int(hex_color[0:2], 16) / 255.0
                            g = int(hex_color[2:4], 16) / 255.0
                            b = int(hex_color[4:6], 16) / 255.0
                            return colors.Color(r, g, b)
                    except:
                        pass
                    return None
                
                def get_alignment(excel_align):
                    """轉換 Excel 對齊方式"""
                    align_map = {
                        'left': 'LEFT',
                        'center': 'CENTER',
                        'right': 'RIGHT',
                        'general': 'LEFT'
                    }
                    return align_map.get(excel_align, 'LEFT')
                
                # 如果有格式資訊，套用每個儲存格的格式
                if formatting:
                    for row_idx, row_formats in enumerate(formatting):
                        for col_idx, cell_format in enumerate(row_formats):
                            cell_pos = (col_idx, row_idx)
                            
                            # 字體大小
                            if cell_format.get('font_size'):
                                font_size = max(6, min(cell_format['font_size'], 20))  # 限制字體大小範圍
                                style_commands.append(('FONTSIZE', cell_pos, cell_pos, font_size))
                            
                            # 字體名稱和粗體
                            font_name = 'ChineseFont' if has_chinese_font else 'Helvetica'
                            if cell_format.get('bold'):
                                if has_chinese_font:
                                    font_name = 'ChineseFont'  # 中文字體處理粗體較複雜
                                else:
                                    font_name = 'Helvetica-Bold'
                            
                            style_commands.append(('FONTNAME', cell_pos, cell_pos, font_name))
                            
                            # 對齊方式
                            alignment = get_alignment(cell_format.get('alignment', 'general'))
                            style_commands.append(('ALIGN', cell_pos, cell_pos, alignment))
                            
                            # 背景顏色
                            if cell_format.get('fill_color'):
                                bg_color = hex_to_color(cell_format['fill_color'])
                                if bg_color:
                                    style_commands.append(('BACKGROUND', cell_pos, cell_pos, bg_color))
                            
                            # 字體顏色
                            if cell_format.get('font_color'):
                                text_color = hex_to_color(cell_format['font_color'])
                                if text_color:
                                    style_commands.append(('TEXTCOLOR', cell_pos, cell_pos, text_color))
                else:
                    # 沒有格式資訊時使用預設樣式
                    if has_chinese_font:
                        style_commands.extend([
                            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
                            ('FONTSIZE', (0, 0), (-1, -1), 10),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ])
                    else:
                        style_commands.extend([
                            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 0), (-1, -1), 10),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ])
                
                table_style = TableStyle(style_commands)
                
                table.setStyle(table_style)
                story.append(table)
                
                # 工作表之間加入適當分隔
                if sheet_idx < len(all_sheets_data) - 1:  # 不是最後一個工作表
                    story.append(Spacer(1, 40))  # 表格間距
            
            # 生成 PDF
            doc.build(story)
            
            # 讀取生成的 PDF 位元組
            with open(temp_pdf_path, 'rb') as pdf_file:
                pdf_bytes = pdf_file.read()
            
            print("Excel 轉 PDF 成功")
            return pdf_bytes
            
        finally:
            # 安全地清理臨時 PDF 檔案
            if temp_pdf_path and os.path.exists(temp_pdf_path):
                try:
                    os.unlink(temp_pdf_path)
                except OSError as e:
                    print(f"警告: 無法刪除臨時 PDF 檔案 {temp_pdf_path}: {e}")
                
    except FileNotFoundError:
        print("錯誤: 找不到檔案")
        return None
    except Exception as e:
        print(f"Excel 轉 PDF 過程中發生錯誤: {str(e)}")
        return None

# 儲存當前選中的條目
current_selected_item = None

# 生成檔案顯示內容的函數  
def generate_file_display_content(file_names):
    """在callback中動態生成檔案顯示內容"""
    print('file_names', file_names, type(file_names))
    if file_names and len(file_names) > 0:
        file_items = []
        for i, file_name in enumerate(file_names):
            file_icon = get_file_icon(file_name)
            file_color = get_file_color(file_name)
            
            file_item = dbc.ListGroupItem([
                html.Div([
                    html.Div([
                        html.I(className=file_icon, style={
                            "fontSize": "20px", 
                            "marginRight": "12px", 
                            "color": file_color
                        }),
                        html.Div([
                            html.H6(file_name, style={
                                "margin": "0", 
                                "color": "#212529", 
                                "fontSize": "14px",
                                "whiteSpace": "nowrap",
                                "overflow": "hidden",
                                "textOverflow": "ellipsis",
                                "maxWidth": "200px"
                            }),
                            html.Small("已存在資料庫", style={"color": "#6c757d", "fontSize": "12px"})
                        ], style={"flex": "1", "minWidth": "0"})
                    ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                    html.Div([
                        dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                 id={"type": "delete-existing-file-btn", "index": i})
                    ])
                ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
            ], style={
                "padding": "12px", 
                "marginBottom": "5px",
                "backgroundColor": "#e7f3ff",  # 淺藍色背景
                "borderColor": "#b3d9ff"
            })
            file_items.append(file_item)
        
        return [
            html.H6(f"資料庫檔案 ({len(file_names)})", 
                   style={"margin": "0 0 15px 0", "color": "#495057", "fontSize": "16px"}),
            dbc.ListGroup(file_items, flush=True)
        ]
    else:
        print("沒有檔案")
        return [html.P("尚未上傳任何檔案", style={
            "color": "#6c757d", 
            "textAlign": "center",
            "marginTop": "50px",
            "fontStyle": "italic"
        })]

# 初始化檔案列表內容變數（layout 使用）
db_content = html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
pending_content = html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})

layout = dbc.Container([
        # Toast 通知
        success_toast("rag", message=""),
        error_toast("rag", message=""),
        warning_toast("rag", message=""),
        # 用戶角色存儲
        dcc.Store(id='user-role-store'),
    
    dcc.Loading(
        id="loading-full-page",
        type="dot",
        children=dbc.Row([
            dbc.Col([
                # 標題 + 新增按鈕
                dbc.Row(
                    justify="between",
                    className="mb-2",
                    children=[
                        dbc.Col(html.H5("知識庫條目"), width="auto", className="d-flex align-items-center"),
                        dbc.Col(dbc.Button("新增條目", id="open-modal", color="primary", size="sm"),
                                width="auto", className="d-flex justify-content-end align-items-center")
                    ]
                ),
                # 條目列表
                dbc.ListGroup(
                    id="client-list",
                    children=[],  # 初始為空，透過 callback 載入
                    style={
                        "backgroundColor": "transparent"
                    }
                )
            ], width=3, style={"display": "none"}),

            dbc.Col([
                html.Div(
                    id="content-area",
                    children=[],  # 初始為空
                    style={
                        "borderRadius": "6px",
                        "padding": "20px",
                        "height": "85vh",
                        "boxShadow": "rgba(0, 0, 0, 0.05) 0px 0px 0px 1px",
                        "overflow": "auto",
                        "display": "flex",
                        "flexDirection": "column"
                    }
                )
            ], width=12)
        ]),
        style={
            "display": "flex",
            "alignItems": "center",
            "justifyContent": "center",
            "minHeight": "85vh"
        }
    ),

    # Modal 彈窗
    dbc.Modal([
        dbc.ModalHeader("新增條目", style={"fontWeight": "bold", "fontSize": "24px"}),
        dbc.ModalBody([
            dbc.Input(id="new-client-name", placeholder="輸入條目標題", type="text")
        ]),
        dbc.ModalFooter([
            dbc.Button("取消", id="close-modal", color="secondary", className="me-2"),
            dbc.Button("新增", id="add-client", color="primary")
        ])
    ], id="modal", is_open=False, centered=True),
    
    # 刪除確認 Modal
    dbc.Modal([
        dbc.ModalHeader("確認刪除", style={"fontWeight": "bold", "fontSize": "24px", "color": "#dc3545"}),
        dbc.ModalBody([
            html.P("此操作無法復原，確定要刪除這個知識庫條目嗎？", style={"marginBottom": "0"})
        ]),
        dbc.ModalFooter([
            dbc.Button("取消", id="cancel-delete-modal", color="secondary", className="me-2"),
            dbc.Button("確認刪除", id="confirm-delete-modal", color="danger")
        ])
    ], id="delete-modal", is_open=False, centered=True)
], fluid=True)

from .common import *

# 頁面載入時初始化右側內容區域
@app.callback(
    Output("content-area", "children", allow_duplicate=True),
    Input("url", "pathname"),
    prevent_initial_call='initial_duplicate'
)
def initialize_content_area(pathname):
    # 當進入 RAG 頁面時，顯示預設訊息
    if pathname and 'rag' in pathname.lower():
        return html.Div([
            html.Div([
                html.H4("請選擇左側的知識庫條目", style={
                    "textAlign": "center",
                    "color": "#6c757d",
                    "marginTop": "50px"
                })
            ], style={"flex": "1", "display": "flex", "alignItems": "center", "justifyContent": "center"})
        ], style={"height": "100%", "display": "flex", "flexDirection": "column"})
    return []

# 頁面載入時自動載入條目列表
@app.callback(
    Output("client-list", "children", allow_duplicate=True),
    Input("url", "pathname"),  # 當頁面路徑改變時觸發
    prevent_initial_call='initial_duplicate'  # 允許重複輸出的初始調用
)
def load_client_list(pathname):

    # 檢查是否為 RAG 頁面或初始載入
    if not pathname or pathname == "/" or (pathname and 'rag' in pathname.lower()):
        items = load_initial_items()

        # 過濾掉要隱藏的條目
        items = [item for item in items if item != "4月988價格表"]

        result = [
            dbc.ListGroupItem(
                name,
                id={"type": "client-item", "index": name},
                n_clicks=0,
                style={
                    "cursor": "pointer",
                    "backgroundColor": "white",
                    "border": "1px solid #e0e6ed",
                    "marginBottom": "8px",
                    "borderRadius": "12px",
                    "boxShadow": "rgb(204, 219, 232) 3px 3px 6px 0px inset, rgba(255, 255, 255, 0.5) -3px -3px 6px 1px inset",
                    "fontSize": "1.2rem",
                    "color": "#000000",
                    "fontWeight": "500"
                }
            )
            for name in items
        ]
        return result
    return []

@app.callback(
    Output("modal", "is_open"),
    Output("new-client-name", "value"),
    Input("open-modal", "n_clicks"),
    Input("close-modal", "n_clicks"),
    Input("add-client", "n_clicks"),
    State("modal", "is_open"),
    prevent_initial_call=True
)
def toggle_modal(open_click, close_click, add_click, is_open):
    triggered = callback_context.triggered_id
    if triggered in ["open-modal", "close-modal", "add-client"]:
        return not is_open, ""
    return is_open, ""

# 新增客戶到 ListGroup
@app.callback(
    Output("client-list", "children"),
    Output('rag-error-toast', 'is_open', allow_duplicate=True),
    Output('rag-error-toast', 'children', allow_duplicate=True),
    Output('rag-warning-toast', 'is_open', allow_duplicate=True),
    Output('rag-warning-toast', 'children', allow_duplicate=True),
    Input("add-client", "n_clicks"),
    State("new-client-name", "value"),
    State("client-list", "children"),
    State("user-role-store", "data"),
    prevent_initial_call=True
)
def add_client(n_clicks, new_name, current_list, user_role):
    if not new_name:
        raise PreventUpdate
    
    try:
        # 準備API請求數據
        knowledge_data = {
            "title": new_name,
            "text_content": "",
            "files": None
        }
        
        # 呼叫API在資料庫新增記錄
        import requests
        knowledge_data["user_role"] = user_role or "viewer"
        response = requests.put("http://127.0.0.1:8000/rag/save_knowledge", json=knowledge_data)
        
        if response.status_code == 200:
            # 資料庫新增成功，更新UI
            new_item = dbc.ListGroupItem(
                new_name, 
                id={"type": "client-item", "index": new_name}, 
                n_clicks=0,
                style={
                    "cursor": "pointer",
                    "backgroundColor": "white",
                    "border": "1px solid #e0e6ed",
                    "marginBottom": "8px",
                    "borderRadius": "12px",
                    "boxShadow": "rgb(204, 219, 232) 3px 3px 6px 0px inset, rgba(255, 255, 255, 0.5) -3px -3px 6px 1px inset",
                    "fontSize": "16px",
                    "color": "#000000",
                    "fontWeight": "500"
                }
            )
            current_list.append(new_item)
            return current_list, False, "", False, ""
        elif response.status_code == 403:
            return current_list, False, "", True, "權限不足：僅限編輯者使用此功能"
        else:
            try:
                error_msg = response.json().get('detail', '新增失敗')
            except:
                error_msg = f"HTTP {response.status_code}"
            return current_list, True, f"新增失敗：{error_msg}", False, ""
            
    except Exception as e:
        print(f"[ERROR] 新增條目失敗: {e}")
        return current_list, True, f"新增條目失敗：{str(e)}", False, ""

# 儲存已上傳檔案的變數（模擬全局狀態）
uploaded_files_store = []

MAX_UPLOAD_SIZE_MB = get_env_int('MAX_UPLOAD_SIZE_MB', 50)
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024


def normalize_upload_timestamp(raw_timestamp):
    """將瀏覽器提供的 last_modified 格式安全轉換為秒，允許 None 或毫秒值"""
    try:
        if raw_timestamp is None:
            return datetime.datetime.now().timestamp()

        if isinstance(raw_timestamp, str):
            if raw_timestamp.isdigit():
                raw_timestamp = float(raw_timestamp)
            else:
                # 嘗試解析 ISO 格式字串
                try:
                    return datetime.datetime.fromisoformat(raw_timestamp).timestamp()
                except ValueError:
                    return datetime.datetime.now().timestamp()

        if isinstance(raw_timestamp, (int, float)):
            value = float(raw_timestamp)
            # html input lastModified 通常是毫秒
            if value > 1e11:
                value /= 1000.0
            return value
    except Exception:
        pass

    return datetime.datetime.now().timestamp()

# 檔案類型檢查和圖示判斷函數
def get_file_icon(filename):
    """根據檔案副檔名返回對應的 Font Awesome 圖示類別"""
    extension = filename.lower().split('.')[-1] if '.' in filename else ''
    
    if extension == 'pdf':
        return "fas fa-file-pdf"
    elif extension in ['doc', 'docx']:
        return "fas fa-file-word"
    elif extension in ['xls', 'xlsx']:
        return "fas fa-file-excel"
    else:
        return "fas fa-file"

def get_file_color(filename):
    """根據檔案類型返回對應的顏色"""
    extension = filename.lower().split('.')[-1] if '.' in filename else ''
    
    if extension == 'pdf':
        return "#dc3545"
    elif extension in ['doc', 'docx']:
        return "#2b5ce6"
    elif extension in ['xls', 'xlsx']:
        return "#107c41"
    else:
        return "#495057"

def is_allowed_file(filename):
    """檢查檔案格式是否被允許"""
    if not filename or '.' not in filename:
        return False
    
    extension = filename.lower().split('.')[-1]
    allowed_extensions = ['pdf']
    return extension in allowed_extensions

# 處理檔案上傳
@app.callback(
    Output("database-files-list", "children"),
    Output("pending-files-list", "children"),
    Output('rag-error-toast', 'is_open'),
    Output('rag-error-toast', 'children'),
    Input("upload-data", "contents"),
    State("upload-data", "filename"),
    State("upload-data", "last_modified"),
    State("title-input", "value"),
    prevent_initial_call=True
)
def update_output(list_of_contents, list_of_names, list_of_dates, current_title):
    global uploaded_files_store
    
    # 如果沒有檔案上傳，不要更新顯示
    if list_of_contents is None:
        raise PreventUpdate

    # dash 可能在單檔時傳回字串，統一轉為列表
    if isinstance(list_of_contents, str):
        list_of_contents = [list_of_contents]
    if isinstance(list_of_names, str):
        list_of_names = [list_of_names]
    if list_of_dates is None:
        list_of_dates = [None] * len(list_of_contents)
    elif isinstance(list_of_dates, (int, float, str)):
        list_of_dates = [list_of_dates]

    error_message = ""
    show_error = False
    
    # 獲取現有資料庫檔案
    existing_db_files = []
    if current_title:
        try:
            import requests
            response = requests.get(f"http://127.0.0.1:8000/get_rag_content/{current_title}")
            if response.status_code == 200:
                content_data = response.json()
                existing_db_files = content_data.get('file_names', [])
        except Exception as e:
            print(f"[ERROR] 載入現有檔案失敗: {e}")
    
    if list_of_contents is not None:
        invalid_files = []
        for i, (contents, filename, date) in enumerate(zip(list_of_contents, list_of_names, list_of_dates)):
            if not is_allowed_file(filename):
                invalid_files.append(filename)
                continue

            # 處理 Excel 檔案：轉換為 PDF
            processed_contents = contents
            processed_filename = filename
            upload_timestamp = normalize_upload_timestamp(date)

            content_string = contents.split(',', 1)[1] if ',' in contents else contents
            try:
                decoded_content = base64.b64decode(content_string)
            except Exception as decode_error:
                print(f"[ERROR] 檔案 base64 解碼失敗: {decode_error}")
                show_error = True
                error_msg = f"檔案 {filename} 解析失敗，請重新上傳"
                error_message = error_msg if not error_message else f"{error_message}；{error_msg}"
                continue

            if len(decoded_content) > MAX_UPLOAD_SIZE_BYTES:
                show_error = True
                size_error = f"檔案 {filename} 超過 {MAX_UPLOAD_SIZE_MB}MB 限制，請壓縮或拆分後再上傳"
                error_message = size_error if not error_message else f"{error_message}；{size_error}"
                continue

            # 檢查是否為 Excel 檔案
            extension = filename.lower().split('.')[-1] if '.' in filename else ''
            if extension in ['xls', 'xlsx']:
                try:
                    # 解碼 base64 內容
                    # 轉換 Excel 為 PDF
                    print(f"正在將 {filename} 轉換為 PDF...")
                    print(f"原始檔案大小: {len(decoded_content)} bytes")

                    pdf_bytes = excel_to_pdf_bytes_win32com(decoded_content, filename)
                    
                    if pdf_bytes:
                        # 成功轉換，更新檔案名稱和內容
                        base_name = filename.rsplit('.', 1)[0]  # 移除原始副檔名
                        processed_filename = f"{base_name}.pdf"
                        conversion_success = True
                        
                        print(f"轉換後 PDF 大小: {len(pdf_bytes)} bytes")
                        
                        # 重新編碼為 base64
                        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
                        processed_contents = "data:application/pdf;base64," + pdf_base64
                        
                        # 驗證 base64 編碼是否正確
                        import hashlib
                        test_decode = base64.b64decode(pdf_base64)
                        decode_hash = hashlib.md5(test_decode).hexdigest()[:8]
                        print(f"[驗證] base64 編碼後重新解碼的雜湊: {decode_hash}")
                        
                        # 加上檔案內容的簡短雜湊來確認唯一性
                        content_hash = hashlib.md5(pdf_bytes).hexdigest()[:8]
                        print(f"Excel 檔案 {filename} 已成功轉換為 {processed_filename}，內容雜湊: {content_hash}")
                    else:
                        # 轉換失敗，記錄錯誤但不中斷處理
                        print(f"警告: Excel 檔案 {filename} 轉換失敗，將使用原始檔案")
                        conversion_success = False
                        
                except Exception as e:
                    print(f"處理 Excel 檔案 {filename} 時發生錯誤: {e}")
                    # 轉換失敗時使用原始檔案
                    conversion_success = False
            else:
                conversion_success = False
            
            # 檢查是否已存在相同檔名，如果存在就更新，否則新增
            existing_file_index = next((i for i, f in enumerate(uploaded_files_store) if f['filename'] == processed_filename), -1)
            if existing_file_index >= 0:
                # 更新現有檔案
                uploaded_files_store[existing_file_index] = {
                    'filename': processed_filename,
                    'date': upload_timestamp,
                    'contents': processed_contents,
                    'frontend_converted': conversion_success
                }
            else:
                # 新增檔案
                uploaded_files_store.append({
                    'filename': processed_filename,
                    'date': upload_timestamp,
                    'contents': processed_contents,
                    'frontend_converted': conversion_success
                })
        
        if invalid_files:
            show_error = True
            if len(invalid_files) == 1:
                invalid_msg = "僅支援 .pdf 格式"
            else:
                invalid_msg = f"以下檔案格式不符合要求：{', '.join(invalid_files)}，僅支援 .pdf 格式"
            error_message = invalid_msg if not error_message else f"{error_message}；{invalid_msg}"
    
    # 生成檔案列表顯示 - 同時顯示資料庫檔案和新上傳檔案
    all_file_items = []
    
    # 1. 顯示資料庫中的檔案 (淺藍色背景)
    if existing_db_files:
        for i, file_name in enumerate(existing_db_files):
            file_icon = get_file_icon(file_name)
            file_color = get_file_color(file_name)
            
            file_item = dbc.ListGroupItem([
                html.Div([
                    html.Div([
                        html.I(className=file_icon, style={
                            "fontSize": "20px", 
                            "marginRight": "12px", 
                            "color": file_color
                        }),
                        html.Div([
                            html.H6(file_name, style={
                                "margin": "0", 
                                "color": "#212529", 
                                "fontSize": "14px",
                                "whiteSpace": "nowrap",
                                "overflow": "hidden",
                                "textOverflow": "ellipsis",
                                "maxWidth": "200px"
                            }),
                            html.Small("已存在資料庫", style={"color": "#6c757d", "fontSize": "12px"})
                        ], style={"flex": "1", "minWidth": "0"})
                    ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                    html.Div([
                        dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                 id={"type": "delete-existing-file-btn", "index": i})
                    ])
                ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
            ], style={
                "padding": "12px", 
                "marginBottom": "5px",
                "backgroundColor": "#e7f3ff",  # 淺藍色背景
                "borderColor": "#b3d9ff"
            })
            all_file_items.append(file_item)
    
    # 2. 顯示新上傳的檔案 (淺綠色背景)
    if uploaded_files_store:
        for i, file_info in enumerate(uploaded_files_store):
            file_icon = get_file_icon(file_info['filename'])
            file_color = get_file_color(file_info['filename'])
            
            file_item = dbc.ListGroupItem([
                html.Div([
                    html.Div([
                        html.I(className=file_icon, style={
                            "fontSize": "20px", 
                            "marginRight": "12px", 
                            "color": file_color
                        }),
                        html.Div([
                            html.H6(file_info['filename'], style={"margin": "0", "color": "#212529", "fontSize": "14px"}),
                            html.Small(f"新上傳 - {datetime.datetime.fromtimestamp(file_info['date']).strftime('%Y-%m-%d %H:%M:%S')}", 
                                     style={"color": "#6c757d", "fontSize": "12px"})
                        ], style={"flex": "1"})
                    ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                    html.Div([
                        dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                 id={"type": "delete-file-btn", "index": i})
                    ])
                ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
            ], style={
                "padding": "12px", 
                "marginBottom": "5px",
                "backgroundColor": "#e8f5e8",  # 淺綠色背景
                "borderColor": "#c3e6c3"
            })
            all_file_items.append(file_item)
    
    # 生成最終顯示內容
    if all_file_items:
        total_db_files = len(existing_db_files) if existing_db_files else 0
        total_new_files = len(uploaded_files_store)
        
        file_list_content = html.Div([
            html.Div([
                html.H6(f"檔案列表 (資料庫:{total_db_files} | 新上傳:{total_new_files})", 
                       style={"margin": "0 0 15px 0", "color": "#495057", "fontSize": "16px"})
            ]),
            dbc.ListGroup(all_file_items, flush=True)
        ])
    else:
        file_list_content = html.P("尚未上傳任何檔案", style={
            "color": "#6c757d", 
            "textAlign": "center",
            "marginTop": "50px",
            "fontStyle": "italic"
        })
    
    # Split content into database files and pending files
    database_files_content = []
    pending_files_content = []
    
    if existing_db_files:
        for i, file_name in enumerate(existing_db_files):
            file_icon = get_file_icon(file_name)
            file_color = get_file_color(file_name)
            
            file_item = dbc.ListGroupItem([
                html.Div([
                    html.Div([
                        html.I(className=file_icon, style={
                            "fontSize": "16px", 
                            "marginRight": "8px", 
                            "color": file_color
                        }),
                        html.Div([
                            html.H6(file_name, style={
                                "margin": "0", 
                                "color": "#212529", 
                                "fontSize": "12px",
                                "whiteSpace": "nowrap",
                                "overflow": "hidden",
                                "textOverflow": "ellipsis"
                            })
                        ], style={"flex": "1", "minWidth": "0"})
                    ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                    html.Div([
                        dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                 id={"type": "delete-existing-file-btn", "index": i})
                    ])
                ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
            ], style={"padding": "8px", "marginBottom": "3px"})
            database_files_content.append(file_item)
    
    if uploaded_files_store:
        for i, file_info in enumerate(uploaded_files_store):
            file_icon = get_file_icon(file_info['filename'])
            file_color = get_file_color(file_info['filename'])
            
            file_item = dbc.ListGroupItem([
                html.Div([
                    html.Div([
                        html.I(className=file_icon, style={
                            "fontSize": "16px", 
                            "marginRight": "8px", 
                            "color": file_color
                        }),
                        html.Div([
                            html.H6(file_info['filename'], style={
                                "margin": "0", 
                                "color": "#212529", 
                                "fontSize": "12px"
                            })
                        ], style={"flex": "1"})
                    ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                    html.Div([
                        dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                 id={"type": "delete-file-btn", "index": i})
                    ])
                ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
            ], style={"padding": "8px", "marginBottom": "3px"})
            pending_files_content.append(file_item)
    
    # Return content for both sections
    db_content = dbc.ListGroup(database_files_content, flush=True) if database_files_content else html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
    pending_content = dbc.ListGroup(pending_files_content, flush=True) if pending_files_content else html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
    
    return db_content, pending_content, show_error, error_message

# 處理刪除檔案功能
@app.callback(
    Output("pending-files-list", "children", allow_duplicate=True),
    Input({"type": "delete-file-btn", "index": ALL}, "n_clicks"),
    prevent_initial_call=True
)
def delete_file(n_clicks_list):
    global uploaded_files_store
    
    ctx = callback_context
    if not ctx.triggered or not any(n_clicks_list):
        raise PreventUpdate
    
    # 找出被點擊的刪除按鈕
    triggered_button = ctx.triggered[0]
    if triggered_button["value"]:
        # 找出被點擊按鈕的索引
        for i, n_clicks in enumerate(n_clicks_list):
            if n_clicks and n_clicks > 0:
                # 刪除對應索引的檔案
                if 0 <= i < len(uploaded_files_store):
                    uploaded_files_store.pop(i)
                break
        
        # 重新生成檔案列表
        if uploaded_files_store:
            file_items = []
            for file_info in uploaded_files_store:
                file_icon = get_file_icon(file_info['filename'])
                file_color = get_file_color(file_info['filename'])
                
                file_item = dbc.ListGroupItem([
                    html.Div([
                        html.Div([
                            html.I(className=file_icon, style={
                                "fontSize": "20px", 
                                "marginRight": "12px", 
                                "color": file_color
                            }),
                            html.Div([
                                html.H6(file_info['filename'], style={"margin": "0", "color": "#212529", "fontSize": "14px"}),
                                html.Small(f"上傳時間: {datetime.datetime.fromtimestamp(file_info['date']).strftime('%Y-%m-%d %H:%M:%S')}", 
                                         style={"color": "#6c757d"})
                            ], style={"flex": "1"})
                        ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                        html.Div([
                            dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                     id={"type": "delete-file-btn", "index": len(file_items)})
                        ])
                    ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
                ], style={"padding": "12px", "marginBottom": "5px"})
                file_items.append(file_item)
            
            return html.Div([
                html.Div([
                    html.H6(f"已上傳檔案 ({len(uploaded_files_store)})", 
                           style={"margin": "0 0 15px 0", "color": "#495057", "fontSize": "16px"})
                ]),
                dbc.ListGroup(file_items, flush=True)
            ])
        else:
            return html.P("尚未上傳任何檔案", style={
                "color": "#6c757d", 
                "textAlign": "center",
                "marginTop": "50px",
                "fontStyle": "italic"
            })
    
    raise PreventUpdate

@app.callback(
    Output("content-area", "children"),
    Output("client-list", "children", allow_duplicate=True),
    Input({"type": "client-item", "index": ALL}, "n_clicks"),
    State({"type": "client-item", "index": ALL}, "id"),
    State("client-list", "children"),
    prevent_initial_call=True
)
def display_client_data(n_clicks_list, id_list, current_list):
    global current_selected_item
    
    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate

    # 取得觸發的具體項目
    triggered_prop_id = ctx.triggered[0]["prop_id"]
    
    # 解析出被點擊的項目索引
    if triggered_prop_id != ".":
        import json
        triggered_id = json.loads(triggered_prop_id.split('.')[0])
        client_name = triggered_id['index']
        current_selected_item = client_name
        
        # 從資料庫載入條目內容
        try:
            import requests
            response = requests.get(f"http://127.0.0.1:8000/get_rag_content/{client_name}")
            if response.status_code == 200:
                content_data = response.json()
                text_content = content_data.get('text_content', '')
                has_file = content_data.get('has_file', False)
                file_names = content_data.get('file_names', [])
            else:
                text_content = ''
                has_file = False
                file_names = []
        except Exception as e:
            print(f"[ERROR] 載入條目內容失敗: {e}")
            text_content = ''
            has_file = False
            file_names = []
        
        # 更新列表項目的樣式，設定選中狀態
        updated_list = []
        for item in current_list:
            item_name = item['props']['id']['index']
            if item_name == client_name:
                # 選中項目 - 使用淺藍色背景
                updated_item = {
                    **item,
                    'props': {
                        **item['props'],
                        'style': {
                            "cursor": "pointer",
                            "backgroundColor": "white",
                            "border": "1px solid #007bff",
                            "marginBottom": "8px",
                            "borderRadius": "12px",
                            "boxShadow": "rgb(204, 219, 232) 3px 3px 6px 0px inset, rgba(255, 255, 255, 0.5) -3px -3px 6px 1px inset",
                            "fontSize": "1.2rem",
                            "color": "#000000",
                            "fontWeight": "500"
                        }
                    }
                }
            else:
                # 未選中項目 - 使用默認樣式
                updated_item = {
                    **item,
                    'props': {
                        **item['props'],
                        'style': {
                            "cursor": "pointer",
                            "backgroundColor": "white",
                            "border": "1px solid #e0e6ed",
                            "marginBottom": "8px",
                            "borderRadius": "12px",
                            "boxShadow": "rgb(204, 219, 232) 3px 3px 6px 0px inset, rgba(255, 255, 255, 0.5) -3px -3px 6px 1px inset",
                            "fontSize": "1.2rem",
                            "color": "#000000",
                            "fontWeight": "500"
                        }
                    }
                }
            updated_list.append(updated_item)
        
        # Generate database files content
        db_files_content = []
        if file_names:
            for i, file_name in enumerate(file_names):
                file_icon = get_file_icon(file_name)
                file_color = get_file_color(file_name)
                
                file_item = dbc.ListGroupItem([
                    html.Div([
                        html.Div([
                            html.I(className=file_icon, style={
                                "fontSize": "16px", 
                                "marginRight": "8px", 
                                "color": file_color
                            }),
                            html.Div([
                                html.H6(file_name, style={
                                    "margin": "0", 
                                    "color": "#212529", 
                                    "fontSize": "12px",
                                    "whiteSpace": "nowrap",
                                    "overflow": "hidden",
                                    "textOverflow": "ellipsis"
                                })
                            ], style={"flex": "1", "minWidth": "0"})
                        ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                        html.Div([
                            dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                     id={"type": "delete-existing-file-btn", "index": i})
                        ])
                    ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
                ], style={"padding": "8px", "marginBottom": "3px"})
                db_files_content.append(file_item)
        
        db_content = dbc.ListGroup(db_files_content, flush=True) if db_files_content else html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
        pending_content = html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
        
        # 返回編輯界面
        return html.Div([
            html.Div([
                html.H4("標題", id="content-title", style={
                    "borderBottom": "1px solid #6c757d",
                    "paddingBottom": "10px",
                    "marginBottom": "10px",
                    "color": "#212529"
                }),
                dbc.Input(
                    id="title-input",
                    placeholder="請輸入標題...",
                    type="text",
                    value=client_name,
                    style={"marginBottom": "20px"}
                ),
                # 隱藏文字內容輸入，保留給後端使用
                html.Div([
                    dbc.Textarea(
                        id="content-input",
                        value=text_content,
                        readOnly=True,
                        style={"display": "none"}
                    )
                ], style={"display": "none"}),
                html.H4("知識庫檔案", style={
                    "borderBottom": "1px solid #6c757d",
                    "paddingBottom": "10px",
                    "marginBottom": "10px",
                    "color": "#212529"
                }),
                # TABS 區域
                dbc.Tabs(
                    id="content-tabs",
                    active_tab="file-tab",
                    children=[
                        dbc.Tab(
                            label="上傳檔案",
                            tab_id="file-tab",
                            children=[
                                html.Div([
                                    dbc.Row([
                                        # 左側：拖曳上傳區域
                                        dbc.Col([
                                            dcc.Upload(
                                                id="upload-data",
                                                children=html.Div([
                                                    html.I(className="fas fa-cloud-upload-alt", style={
                                                        "fontSize": "2.5rem",
                                                        "color": "#007bff",
                                                        "marginBottom": "1rem"
                                                    }),
                                                    html.P("拖拽檔案到此處或點擊上傳", style={
                                                        "fontSize": "1rem",
                                                        "color": "#666",
                                                        "margin": "0"
                                                    }),
                                                    html.P("支援格式：PDF", style={
                                                        "fontSize": "0.8rem",
                                                        "color": "#999",
                                                        "margin": "0.5rem 0 0 0"
                                                    }),
                                                    html.P(
                                                        f"大小限制：≤ {MAX_UPLOAD_SIZE_MB} MB",
                                                        style={
                                                            "fontSize": "0.8rem",
                                                            "color": "#999",
                                                            "margin": "0.2rem 0 0 0"
                                                        }
                                                    ),
                                                ], style={
                                                    "display": "flex",
                                                    "flexDirection": "column",
                                                    "alignItems": "center",
                                                    "justifyContent": "center",
                                                    "height": "100%"
                                                }),
                                                style={
                                                    "width": "100%",
                                                    "height": "45vh",
                                                    "borderWidth": "2px",
                                                    "borderStyle": "dashed",
                                                    "borderColor": "#007bff",
                                                    "borderRadius": "12px",
                                                    "textAlign": "center",
                                                    "backgroundColor": "#f8f9ff",
                                                    "cursor": "pointer",
                                                    "transition": "all 0.3s ease"
                                                },
                                                multiple=True
                                            )
                                        ], width=5),
                                        
                                        # 右側：檔案列表區域
                                        dbc.Col([
                                            # 上半部：資料庫檔案區塊
                                            html.Div([
                                                html.Div([
                                                    html.H6("已上傳檔案", style={
                                                        "margin": "0 0 10px 0", 
                                                        "color": "#495057", 
                                                        "fontSize": "14px",
                                                        "fontWeight": "600"
                                                    }),
                                                    html.Div(
                                                        id="database-files-list",
                                                        children=db_content,
                                                        style={
                                                            "height": "calc(100% - 30px)",
                                                            "overflowY": "auto"
                                                        }
                                                    )
                                                ], style={
                                                    "border": "1px solid #dee2e6",
                                                    "borderRadius": "6px",
                                                    "padding": "10px",
                                                    "backgroundColor": "white",
                                                    "height": "100%",
                                                    "display": "flex",
                                                    "flexDirection": "column"
                                                })
                                            ], style={
                                                "height": "calc(50% - 5px)",
                                                "marginBottom": "10px"
                                            }),
                                            
                                            # 下半部：待上傳檔案區塊
                                            html.Div([
                                                html.Div([
                                                    html.H6("待上傳檔案", style={
                                                        "margin": "0 0 10px 0", 
                                                        "color": "#495057", 
                                                        "fontSize": "14px",
                                                        "fontWeight": "600"
                                                    }),
                                                    html.Div(
                                                        id="pending-files-list",
                                                        children=pending_content,
                                                        style={
                                                            "height": "calc(100% - 30px)",
                                                            "overflowY": "auto"
                                                        }
                                                    )
                                                ], style={
                                                    "border": "1px solid #dee2e6",
                                                    "borderRadius": "6px",
                                                    "padding": "10px",
                                                    "backgroundColor": "white",
                                                    "height": "100%",
                                                    "display": "flex",
                                                    "flexDirection": "column"
                                                })
                                            ], style={
                                                "height": "calc(50% - 5px)"
                                            })
                                        ], width=7, style={
                                            "height": "45vh",
                                            "display": "flex",
                                            "flexDirection": "column"
                                        })
                                    ], className="g-2")
                                ], className="mt-3", style={"maxWidth": "98%", "margin": "0 auto"})
                            ]
                        )
                    ]
                )
            ], style={"flex": "1", "overflowY": "auto"}),
            # 儲存按鈕區域
            html.Div([
                dbc.Button("儲存修改內容", id="save-btn", color="success", size="sm"),
                dbc.Button(f"刪除此條目", id="delete-current-item-btn", color="danger", size="sm",
                          style={"marginLeft": "10px"})
            ], style={
                "display": "flex",
                "justifyContent": "center",
                "alignItems": "center",
                "paddingTop": "10px",
                "marginTop": "10px"
            })
        ], style={"height": "100%", "display": "flex", "flexDirection": "column"}), updated_list

    # 如果沒有點選任何項目，顯示提示訊息
    return html.Div([
        html.Div([
            html.H4("請選擇左側的知識庫條目", style={
                "textAlign": "center",
                "color": "#6c757d",
                "marginTop": "50px"
            })
        ], style={"flex": "1", "display": "flex", "alignItems": "center", "justifyContent": "center"})
    ], style={"height": "100%", "display": "flex", "flexDirection": "column"}), no_update

# 處理刪除此條目按鈕點擊 - 顯示確認Modal
@app.callback(
    Output("delete-modal", "is_open", allow_duplicate=True),
    Input("delete-current-item-btn", "n_clicks"),
    State("delete-modal", "is_open"),
    State("content-area", "children"),
    prevent_initial_call=True
)
def show_delete_confirmation_modal(n_clicks, is_open, current_content):
    global item_to_delete
    
    if not n_clicks:
        raise PreventUpdate
    
    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate
    
    # 從目前的內容區域中找到正在編輯的條目名稱
    current_item_name = None
    try:
        # 尋找title-input的值
        if current_content and 'props' in current_content:
            children = current_content['props'].get('children', [])
            for child in children:
                if isinstance(child, dict) and 'props' in child:
                    grandchildren = child['props'].get('children', [])
                    for grandchild in grandchildren:
                        if isinstance(grandchild, dict) and 'props' in grandchild:
                            if grandchild['props'].get('id') == 'title-input':
                                current_item_name = grandchild['props'].get('value', '')
                                break
    except:
        pass
    
    if not current_item_name:
        raise PreventUpdate
    
    # 儲存要刪除的條目名稱
    item_to_delete = current_item_name
    
    # 顯示確認Modal
    return True

# 處理刪除確認Modal的按鈕點擊
@app.callback(
    Output("delete-modal", "is_open", allow_duplicate=True),
    Output("client-list", "children", allow_duplicate=True),
    Output("content-area", "children", allow_duplicate=True),
    Output('rag-warning-toast', 'is_open', allow_duplicate=True),
    Output('rag-warning-toast', 'children', allow_duplicate=True),
    Input("confirm-delete-modal", "n_clicks"),
    Input("cancel-delete-modal", "n_clicks"),
    State("delete-modal", "is_open"),
    State("client-list", "children"),
    State("user-role-store", "data"),
    prevent_initial_call=True
)
def handle_delete_modal_buttons(confirm_clicks, cancel_clicks, is_open, current_list, user_role):
    global item_to_delete
    
    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate
    
    triggered_button = ctx.triggered[0]['prop_id']
    
    # 如果點擊取消按鈕，關閉Modal
    if 'cancel-delete-modal' in triggered_button:
        item_to_delete = None
        return False, current_list, no_update, False, ""
    
    # 如果點擊確認刪除按鈕
    if 'confirm-delete-modal' in triggered_button and item_to_delete:
        try:
            # 呼叫API從資料庫刪除記錄
            import requests
            response = requests.put(
                f"http://127.0.0.1:8000/rag/delete_knowledge/{item_to_delete}",
                params={"user_role": user_role or "viewer"}
            )
            
            if response.status_code == 200:
                # 資料庫刪除成功，從UI列表中移除選中的條目
                updated_list = []
                for item in current_list:
                    if item['props']['id']['index'] != item_to_delete:
                        updated_list.append(item)
                
                # 重置刪除狀態
                item_to_delete = None
                
                # 返回更新後的列表和默認內容區域
                default_content = html.Div([
                    html.Div([
                        html.H4("請選擇左側的知識庫條目", style={
                            "textAlign": "center",
                            "color": "#6c757d",
                            "marginTop": "50px"
                        })
                    ], style={"flex": "1", "display": "flex", "alignItems": "center", "justifyContent": "center"})
                ], style={"height": "100%", "display": "flex", "flexDirection": "column"})
                
                # 關閉Modal，更新列表，返回默認內容區域
                return False, updated_list, default_content, False, ""
            elif response.status_code == 403:
                # 權限不足，顯示警告並關閉Modal
                item_to_delete = None
                return False, current_list, no_update, True, "權限不足：僅限編輯者使用此功能"
            else:
                # 刪除失敗，保持現狀但關閉Modal
                item_to_delete = None
                return False, current_list, no_update, False, ""
                
        except Exception as e:
            print(f"[ERROR] 刪除條目失敗: {e}")
            # 發生錯誤，保持現狀但關閉Modal
            item_to_delete = None
            return False, current_list, no_update, False, ""
    
    raise PreventUpdate


# 處理儲存按鈕點擊
@app.callback(
    Output('rag-success-toast', 'is_open', allow_duplicate=True),
    Output('rag-success-toast', 'children', allow_duplicate=True),
    Output('rag-error-toast', 'is_open', allow_duplicate=True),
    Output('rag-error-toast', 'children', allow_duplicate=True),
    Output('rag-warning-toast', 'is_open', allow_duplicate=True),
    Output('rag-warning-toast', 'children', allow_duplicate=True),
    Output("database-files-list", "children", allow_duplicate=True),
    Output("pending-files-list", "children", allow_duplicate=True),
    Output("client-list", "children", allow_duplicate=True),
    Input("save-btn", "n_clicks"),
    State("title-input", "value"),
    State("content-input", "value"),
    State("client-list", "children"),
    State("user-role-store", "data"),
    prevent_initial_call=True
)
def handle_save_button(n_clicks, title, text_content, current_list, user_role):
    global uploaded_files_store, current_selected_item
    
    if not n_clicks:
        raise PreventUpdate
    
    if not title:
        return False, "", True, "請輸入標題", False, "", no_update, no_update, no_update
    
    try:
        # 檢查標題是否有變更，如果有變更先更新標題
        title_updated = False
        old_title = current_selected_item  # 記住原標題
        if current_selected_item and title != current_selected_item:
            import requests
            update_data = {
                "old_title": current_selected_item,
                "new_title": title
            }
            update_data["user_role"] = user_role or "viewer"
            response = requests.put("http://127.0.0.1:8000/rag/update_title", json=update_data)
            
            if response.status_code == 200:
                title_updated = True
                current_selected_item = title
            elif response.status_code == 403:
                return False, "", False, "", True, "權限不足：僅限編輯者使用此功能", no_update, no_update, no_update
            else:
                error_msg = response.json().get('detail', '標題更新失敗')
                return False, "", True, f"標題更新失敗：{error_msg}", False, "", no_update, no_update, no_update
        
        # 準備檔案數據
        files_data = []
        if uploaded_files_store:
            for file_info in uploaded_files_store:
                # 提取 base64 內容
                content_base64 = file_info['contents'].split(',')[1] if ',' in file_info['contents'] else file_info['contents']
                
                # 計算內容雜湊用於除錯
                import hashlib
                # 解碼 base64 來計算原始內容的雜湊
                try:
                    decoded_for_hash = base64.b64decode(content_base64)
                    pdf_content_hash = hashlib.md5(decoded_for_hash).hexdigest()[:8]
                    base64_hash = hashlib.md5(content_base64.encode()).hexdigest()[:8]
                    print(f"準備儲存檔案: {file_info['filename']}")
                    print(f"  - PDF 內容雜湊: {pdf_content_hash}")
                    print(f"  - base64 字串雜湊: {base64_hash}")
                except:
                    content_hash = hashlib.md5(content_base64.encode()).hexdigest()[:8]
                    print(f"準備儲存檔案: {file_info['filename']}, 內容雜湊: {content_hash}")
                
                files_data.append({
                    'filename': file_info['filename'],
                    'content': content_base64,
                    'frontend_converted': file_info.get('frontend_converted', False)
                })
        
        # 準備API請求數據
        knowledge_data = {
            "title": title,
            "text_content": text_content or "",
            "files": files_data if files_data else None
        }
        
        # 呼叫API儲存內容
        import requests
        knowledge_data["user_role"] = user_role or "viewer"
        response = requests.put("http://127.0.0.1:8000/rag/save_knowledge", json=knowledge_data)
        
        if response.status_code == 200:
            # 清空上傳檔案暫存（因為已經儲存到資料庫）
            uploaded_files_store.clear()
            
            # 重新從資料庫載入檔案列表
            try:
                content_response = requests.get(f"http://127.0.0.1:8000/get_rag_content/{title}")
                if content_response.status_code == 200:
                    content_data = content_response.json()
                    updated_file_names = content_data.get('file_names', [])
                else:
                    updated_file_names = []
            except:
                updated_file_names = []
            
            # Generate content for database files and clear pending files
            db_files_content = []
            if updated_file_names:
                for i, file_name in enumerate(updated_file_names):
                    file_icon = get_file_icon(file_name)
                    file_color = get_file_color(file_name)
                    
                    file_item = dbc.ListGroupItem([
                        html.Div([
                            html.Div([
                                html.I(className=file_icon, style={
                                    "fontSize": "16px", 
                                    "marginRight": "8px", 
                                    "color": file_color
                                }),
                                html.Div([
                                    html.H6(file_name, style={
                                        "margin": "0", 
                                        "color": "#212529", 
                                        "fontSize": "12px",
                                        "whiteSpace": "nowrap",
                                        "overflow": "hidden",
                                        "textOverflow": "ellipsis"
                                    })
                                ], style={"flex": "1", "minWidth": "0"})
                            ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                            html.Div([
                                dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                         id={"type": "delete-existing-file-btn", "index": i})
                            ])
                        ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
                    ], style={"padding": "8px", "marginBottom": "3px"})
                    db_files_content.append(file_item)
            
            db_content = dbc.ListGroup(db_files_content, flush=True) if db_files_content else html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
            pending_content = html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
            
            # 更新左側列表如果標題有變更
            updated_list = current_list
            if title_updated and old_title:
                updated_list = []
                for item in current_list:
                    if 'props' in item and 'id' in item['props']:
                        item_index = item['props']['id'].get('index')
                        if item_index == old_title:
                            # 更新選中項目的標題
                            updated_item = {
                                **item,
                                'props': {
                                    **item['props'],
                                    'children': title,
                                    'id': {"type": "client-item", "index": title},
                                    'style': {
                                        "cursor": "pointer",
                                        "backgroundColor": "white",
                                        "border": "1px solid #007bff",
                                        "marginBottom": "8px",
                                        "borderRadius": "12px",
                                        "boxShadow": "rgb(204, 219, 232) 3px 3px 6px 0px inset, rgba(255, 255, 255, 0.5) -3px -3px 6px 1px inset",
                                        "fontSize": "1.2rem",
                                        "color": "#000000",
                                        "fontWeight": "500"
                                    }
                                }
                            }
                            updated_list.append(updated_item)
                        else:
                            updated_list.append(item)
                    else:
                        updated_list.append(item)
            
            success_msg = "知識庫內容儲存成功！"
            if title_updated:
                success_msg = "標題和內容儲存成功！"
            
            return True, success_msg, False, "", False, "", db_content, pending_content, updated_list
        elif response.status_code == 403:
            return False, "", False, "", True, "權限不足：僅限編輯者使用此功能", no_update, no_update, no_update
        else:
            error_msg = response.json().get('detail', '儲存失敗')
            return False, "", True, f"儲存失敗：{error_msg}", False, "", no_update, no_update, no_update
            
    except Exception as e:
        print(f"[ERROR] 儲存失敗: {e}")
        return False, "", True, f"儲存失敗：{str(e)}", False, "", no_update, no_update, no_update

# 處理刪除已存在檔案
@app.callback(
    Output('rag-success-toast', 'is_open', allow_duplicate=True),
    Output('rag-success-toast', 'children', allow_duplicate=True),
    Output('rag-warning-toast', 'is_open', allow_duplicate=True),
    Output('rag-warning-toast', 'children', allow_duplicate=True),
    Output("database-files-list", "children", allow_duplicate=True),
    Input({"type": "delete-existing-file-btn", "index": ALL}, "n_clicks"),
    State("title-input", "value"),
    State("content-input", "value"),
    State("user-role-store", "data"),
    prevent_initial_call=True
)
def delete_existing_file(n_clicks_list, title, text_content, user_role):
    ctx = callback_context
    if not ctx.triggered or not any(n_clicks_list) or not title:
        raise PreventUpdate
    
    # 找出被點擊的按鈕索引
    triggered_button = ctx.triggered[0]
    if not triggered_button["value"]:
        raise PreventUpdate
        
    # 解析出要刪除的檔案索引
    import json
    button_id = json.loads(triggered_button["prop_id"].split('.')[0])
    file_index = button_id["index"]
    
    try:
        # 呼叫API清除檔案內容
        import requests
        
        # 準備刪除指定檔案的數據
        knowledge_data = {
            "title": title,
            "text_content": text_content or "",
            "files": None,
            "delete_file_index": file_index
        }
        
        # 更新資料庫，清除file_content和file_name
        knowledge_data["user_role"] = user_role or "viewer"
        response = requests.put("http://127.0.0.1:8000/rag/save_knowledge", json=knowledge_data)
        
        if response.status_code == 200:
            # 重新載入檔案列表
            try:
                content_response = requests.get(f"http://127.0.0.1:8000/get_rag_content/{title}")
                if content_response.status_code == 200:
                    content_data = content_response.json()
                    updated_file_names = content_data.get('file_names', [])
                    # Generate updated database files content
                    db_files_content = []
                    if updated_file_names:
                        for i, file_name in enumerate(updated_file_names):
                            file_icon = get_file_icon(file_name)
                            file_color = get_file_color(file_name)
                            
                            file_item = dbc.ListGroupItem([
                                html.Div([
                                    html.Div([
                                        html.I(className=file_icon, style={
                                            "fontSize": "16px", 
                                            "marginRight": "8px", 
                                            "color": file_color
                                        }),
                                        html.Div([
                                            html.H6(file_name, style={
                                                "margin": "0", 
                                                "color": "#212529", 
                                                "fontSize": "12px",
                                                "whiteSpace": "nowrap",
                                                "overflow": "hidden",
                                                "textOverflow": "ellipsis"
                                            })
                                        ], style={"flex": "1", "minWidth": "0"})
                                    ], style={"display": "flex", "alignItems": "center", "flex": "1"}),
                                    html.Div([
                                        dbc.Button("刪除", color="danger", size="sm", outline=True, 
                                                 id={"type": "delete-existing-file-btn", "index": i})
                                    ])
                                ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center"})
                            ], style={"padding": "8px", "marginBottom": "3px"})
                            db_files_content.append(file_item)
                    
                    db_content = dbc.ListGroup(db_files_content, flush=True) if db_files_content else html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
                    return True, "檔案刪除成功！", False, "", db_content
                else:
                    db_content = html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
                return True, "檔案刪除成功！", False, "", db_content
            except:
                db_content = html.P("無檔案", style={"color": "#6c757d", "textAlign": "center", "margin": "20px 0"})
                return True, "檔案刪除成功！", False, "", db_content
        elif response.status_code == 403:
            return False, "", True, "權限不足：僅限編輯者使用此功能", dash.no_update
        else:
            # 刪除失敗，保持現狀
            raise PreventUpdate
            
    except Exception as e:
        print(f"[ERROR] 刪除檔案失敗: {e}")
        raise PreventUpdate
