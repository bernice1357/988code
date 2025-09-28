# TODO 要改回ORDER_TRANSACTIONS

from fastapi import APIRouter, UploadFile, File, HTTPException, Form
import sys
import os
# 新增資料庫連線管理
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database_config import get_db_connection, execute_query, execute_transaction
from env_loader import load_env_file

# 載入環境變數
load_env_file()
from fastapi.responses import JSONResponse
import datetime
import tempfile
import os
import pandas as pd
import random
import string
import openpyxl
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import RealDictCursor
import logging
from typing import List, Dict, Optional, Tuple

# 設置日誌
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter()

def check_editor_permission(user_role: str):
    """檢查是否有編輯權限"""
    if user_role != 'editor':
        raise HTTPException(status_code=403, detail="權限不足：僅限編輯者使用此功能")

# 備用配置
DEFAULT_CONFIG = {
    'batch_size': 100,
    'timeout': 30
}

class SalesDataUploader:
    def __init__(self):
        """初始化數據上傳器"""
        self.connection = None
        
        # 這個模組專用的表配置
        self.table_config = {
            'order_transactions': 'order_transactions',
            'product_master': 'product_master',
            'customer': 'customer',  # 新增客戶表配置
            'inventory': 'inventory'  # 新增庫存表配置
        }
        
    def connect_database(self):
        """連接數據庫"""
        try:
            # 使用新的統一資料庫連接管理
            db_manager = get_db_connection()
            self.connection = db_manager.__enter__()
            self.db_manager = db_manager  # 保存 context manager 以便後續關閉
            logger.info("數據庫連接成功")
            return True
        except Exception as e:
            logger.error(f"數據庫連接失敗: {str(e)}")
            return False
    
    def close_connection(self):
        """關閉數據庫連接"""
        if hasattr(self, 'db_manager') and self.db_manager:
            self.db_manager.__exit__(None, None, None)
            logger.info("數據庫連接已關閉")
    
    def delete_records_by_months(self, months: set) -> int:
        """
        刪除多個年月的記錄
        
        Args:
            months (set): 包含 (年份, 月份) 元組的集合
            
        Returns:
            int: 刪除的記錄總數
        """
        if not months:
            return 0
            
        total_deleted = 0
        
        try:
            for year, month in months:
                with self.connection.cursor() as cursor:
                    query = f"""
                    DELETE FROM {self.table_config['order_transactions']}
                    WHERE EXTRACT(YEAR FROM transaction_date) = %s 
                      AND EXTRACT(MONTH FROM transaction_date) = %s
                    """
                    cursor.execute(query, (year, month))
                    deleted_count = cursor.rowcount
                    total_deleted += deleted_count
                    
                    logger.info(f"刪除了 {deleted_count} 筆 {year}年{month}月 的記錄")
            
            self.connection.commit()
            logger.info(f"總共刪除了 {total_deleted} 筆記錄")
            return total_deleted
                
        except Exception as e:
            self.connection.rollback()
            logger.error(f"刪除月份記錄失敗: {str(e)}")
            raise
    
    def extract_months_from_data(self, data: List[Dict]) -> set:
        """
        從數據中提取所有涉及的年月
        
        Args:
            data (list): 交易記錄列表
            
        Returns:
            set: 包含 (年份, 月份) 元組的集合
        """
        months = set()
        
        for record in data:
            transaction_date = record['transaction_date']
            if transaction_date:
                # 處理不同的日期格式
                if isinstance(transaction_date, str):
                    # 如果是字符串，嘗試解析
                    try:
                        date_obj = datetime.datetime.strptime(transaction_date, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            date_obj = datetime.datetime.strptime(transaction_date, '%Y/%m/%d').date()
                        except ValueError:
                            logger.warning(f"無法解析日期格式: {transaction_date}")
                            continue
                elif hasattr(transaction_date, 'year') and hasattr(transaction_date, 'month'):
                    # 如果已經是日期對象
                    date_obj = transaction_date
                else:
                    logger.warning(f"未知日期格式: {transaction_date}")
                    continue
                
                months.add((date_obj.year, date_obj.month))
        
        return months
    
    def generate_unique_transaction_id(self) -> str:
        """
        生成唯一的 8 位數字+小寫英文 transaction_id
        格式: 8位隨機字符 (數字+小寫字母) 例如: a1b2c3d4
        
        Returns:
            str: 唯一的 transaction_id
        """
        max_attempts = 50  # 增加嘗試次數
        for attempt in range(max_attempts):
            # 使用更強的隨機性：時間戳 + 隨機數 + 計數器
            import time
            current_time = time.time_ns()  # 納秒級時間戳
            timestamp_part = str(current_time)[-4:]  # 取後4位
            random_part = ''.join(random.choices(string.ascii_lowercase + string.digits, k=4))
            transaction_id = timestamp_part + random_part
            
            # 檢查是否已存在（僅在有資料庫連接時）
            if self.connection and not self.transaction_id_exists(transaction_id):
                return transaction_id
            elif not self.connection:
                # 如果沒有資料庫連接，直接返回（在解析階段）
                return transaction_id
            
            # 如果重複，等待一小段時間增加隨機性
            time.sleep(0.001)
        
        # 如果多次嘗試都失敗，使用完全隨機的ID + UUID後綴
        import uuid
        random_id = ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))
        uuid_suffix = str(uuid.uuid4())[:2]
        return random_id + uuid_suffix
    
    def transaction_id_exists(self, transaction_id: str) -> bool:
        """
        檢查 transaction_id 是否已存在
        
        Args:
            transaction_id (str): 要檢查的 transaction_id
            
        Returns:
            bool: 如果已存在返回 True，否則返回 False
        """
        try:
            with self.connection.cursor() as cursor:
                query = f"""
                SELECT EXISTS(
                    SELECT 1 FROM {self.table_config['order_transactions']}
                    WHERE transaction_id = %s
                )
                """
                cursor.execute(query, (transaction_id,))
                return cursor.fetchone()[0]
                
        except Exception as e:
            logger.error(f"檢查 transaction_id 存在性失敗: {str(e)}")
            # 回滾當前事務以清除錯誤狀態
            try:
                self.connection.rollback()
            except:
                pass
            return True  # 出錯時保守處理，假設已存在

    def check_customer_exists(self, customer_id: str) -> bool:
        """
        檢查客戶是否已存在
        
        Args:
            customer_id (str): 客戶ID
            
        Returns:
            bool: 如果已存在返回 True，否則返回 False
        """
        try:
            with self.connection.cursor() as cursor:
                query = f"""
                SELECT EXISTS(
                    SELECT 1 FROM {self.table_config['customer']}
                    WHERE customer_id = %s
                )
                """
                cursor.execute(query, (customer_id,))
                return cursor.fetchone()[0]
                
        except Exception as e:
            logger.error(f"檢查客戶存在性失敗: {str(e)}")
            # 回滾當前事務以清除錯誤狀態
            try:
                self.connection.rollback()
            except:
                pass
            return False  # 出錯時假設不存在
    
    def check_product_exists(self, product_id: str) -> bool:
        """
        檢查產品是否已存在於 product_master 表中
        
        Args:
            product_id (str): 產品ID
            
        Returns:
            bool: 如果已存在返回 True，否則返回 False
        """
        try:
            with self.connection.cursor() as cursor:
                query = f"""
                SELECT EXISTS(
                    SELECT 1 FROM {self.table_config['product_master']}
                    WHERE product_id = %s
                )
                """
                cursor.execute(query, (product_id,))
                return cursor.fetchone()[0]
                
        except Exception as e:
            logger.error(f"檢查產品存在性失敗: {str(e)}")
            try:
                self.connection.rollback()
            except:
                pass
            return False

    def get_missing_products(self, data: List[Dict]) -> List[str]:
        """
        檢查數據中哪些產品不存在於 product_master 表中
        
        Args:
            data (list): 交易記錄列表
            
        Returns:
            list: 不存在的產品ID列表
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        
        # 提取所有唯一的產品ID
        unique_product_ids = set()
        for record in data:
            product_id = record.get('product_id')
            if product_id and product_id.strip():
                unique_product_ids.add(product_id.strip())
        
        missing_products = []
        
        for product_id in unique_product_ids:
            if not self.check_product_exists(product_id):
                missing_products.append(product_id)
                logger.info(f"發現新產品ID: {product_id}")
        
        return missing_products

    def filter_valid_records(self, data: List[Dict]) -> Tuple[List[Dict], int, List[str], List[str]]:
        """
        過濾出有效的記錄（客戶和產品都存在的記錄）
        
        Args:
            data (list): 原始交易記錄列表
            
        Returns:
            tuple: (有效記錄列表, 跳過記錄數, 跳過的客戶ID列表, 跳過的產品ID列表)
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        
        valid_records = []
        skipped_customers = set()
        skipped_products = set()
        skipped_count = 0
        
        for record in data:
            customer_id = record.get('customer_id', '').strip()
            product_id = record.get('product_id', '').strip()
            
            # 檢查客戶是否存在
            customer_exists = self.check_customer_exists(customer_id) if customer_id else False
            # 檢查產品是否存在  
            product_exists = self.check_product_exists(product_id) if product_id else False
            
            if customer_exists and product_exists:
                # 客戶和產品都存在，記錄有效
                valid_records.append(record)
            else:
                # 記錄跳過的項目
                skipped_count += 1
                if not customer_exists and customer_id:
                    skipped_customers.add(customer_id)
                if not product_exists and product_id:
                    skipped_products.add(product_id)
                    
                logger.info(f"跳過記錄: customer_id={customer_id}, product_id={product_id}, "
                           f"customer_exists={customer_exists}, product_exists={product_exists}")
        
        logger.info(f"記錄過濾完成: 有效={len(valid_records)}, 跳過={skipped_count}")
        return valid_records, skipped_count, list(skipped_customers), list(skipped_products)

    def filter_valid_records_by_customer(self, data: List[Dict]) -> Tuple[List[Dict], int, List[str]]:
        """
        只根據客戶過濾有效記錄
        
        Args:
            data (list): 原始交易記錄列表
            
        Returns:
            tuple: (有效記錄列表, 跳過記錄數, 跳過的客戶ID列表)
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        
        valid_records = []
        skipped_customers = set()
        skipped_count = 0
        
        for record in data:
            customer_id = record.get('customer_id', '').strip()
            
            # 檢查客戶是否存在
            customer_exists = self.check_customer_exists(customer_id) if customer_id else False
            
            if customer_exists:
                # 客戶存在，記錄有效
                valid_records.append(record)
            else:
                # 記錄跳過的客戶
                skipped_count += 1
                if customer_id:
                    skipped_customers.add(customer_id)
                    
                logger.info(f"跳過記錄（缺失客戶）: customer_id={customer_id}")
        
        logger.info(f"客戶記錄過濾完成: 有效={len(valid_records)}, 跳過={skipped_count}")
        return valid_records, skipped_count, list(skipped_customers)

    def process_file_with_product_check(self, file_path: str, delete_month_records: bool = True) -> Tuple[int, int, int, List[str], List[str]]:
        """
        處理整個流程：解析文件 -> 上傳存在的記錄 -> 返回缺失項目列表
        
        Returns:
            tuple: (刪除記錄數, 插入記錄數, 跳過記錄數, 缺失客戶列表, 缺失產品列表)
        """
        deleted_count = 0
        
        try:
            # 連接數據庫
            if not self.connect_database():
                raise Exception("無法連接數據庫")
            
            # 1. 解析 Excel 文件
            logger.info("解析 Excel 文件以提取數據...")
            data = self.parse_sales_data(file_path)
            
            # 2. 過濾有效記錄（混合模式：上傳存在的，返回缺失的）
            valid_records, skipped_count, skipped_customers, skipped_products = self.filter_valid_records(data)
            
            if skipped_count > 0:
                logger.info(f"將上傳 {len(valid_records)} 筆有效記錄，跳過 {skipped_count} 筆記錄（缺失 {len(skipped_customers)} 個客戶和 {len(skipped_products)} 個產品）")
            
            # 3. 上傳有效記錄（即使有跳過的記錄也要上傳）
            if valid_records:
                # 從有效記錄中提取月份資訊進行刪除
                months_to_delete = self.extract_months_from_data(valid_records)
                
                if delete_month_records and months_to_delete:
                    deleted_count = self.delete_records_by_months(months_to_delete)
                
                # 插入有效記錄
                inserted_count = self.insert_all_records(valid_records)
            else:
                inserted_count = 0
            
            # 4. 返回結果：包含上傳統計和缺失項目
            return deleted_count, inserted_count, skipped_count, skipped_customers, skipped_products
            
        finally:
            self.close_connection()

    def get_missing_customers(self, data: List[Dict]) -> List[str]:
        """
        檢查數據中哪些客戶不存在於資料庫中
        
        Args:
            data (list): 交易記錄列表
            
        Returns:
            list: 不存在的客戶ID列表
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        
        # 提取所有唯一的客戶ID
        unique_customer_ids = set()
        for record in data:
            customer_id = record.get('customer_id')
            if customer_id and customer_id.strip():
                unique_customer_ids.add(customer_id.strip())
        
        missing_customers = []
        
        for customer_id in unique_customer_ids:
            if not self.check_customer_exists(customer_id):
                missing_customers.append(customer_id)
                logger.info(f"發現新客戶ID: {customer_id}")
        
        return missing_customers
    
    def parse_sales_data(self, file_path: str) -> List[Dict]:
        """
        解析 Excel 文件，提取交易數據
        
        Args:
            file_path (str): Excel 文件路徑
        
        Returns:
            list: 包含所有交易記錄的列表
        """
        logger.info(f"開始解析文件: {file_path}")
        
        # 使用 openpyxl 讀取 Excel 文件
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        
        logger.info(f"工作表名稱: {worksheet.title}")
        logger.info(f"數據範圍: {worksheet.calculate_dimension()}")
        
        # 獲取最大行數和列數
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        logger.info(f"總行數: {max_row}, 總列數: {max_col}")
        
        # 數據從第8行開始
        data = []
        
        # 用於向前填充的變量
        current_customer_id = ''
        current_product_id = ''
        current_product_name = ''
        current_currency = ''
        
        # 遍歷從第8行開始的所有行
        for row in range(8, max_row + 1):
            # 獲取各列的值
            customer_id_cell = worksheet.cell(row=row, column=1).value    # A列
            product_id_cell = worksheet.cell(row=row, column=6).value     # F列
            product_name_cell = worksheet.cell(row=row, column=11).value  # K列
            transaction_date = worksheet.cell(row=row, column=14).value   # N列
            document_type = worksheet.cell(row=row, column=17).value or '' # Q列
            quantity = worksheet.cell(row=row, column=19).value           # S列
            unit_price = worksheet.cell(row=row, column=22).value or 0    # V列
            currency_cell = worksheet.cell(row=row, column=25).value      # Y列
            amount = worksheet.cell(row=row, column=26).value or 0        # Z列
            
            # 向前填充邏輯
            if customer_id_cell and str(customer_id_cell).strip():
                current_customer_id = str(customer_id_cell).strip()
                
            if product_id_cell and str(product_id_cell).strip():
                current_product_id = str(product_id_cell).strip()
                
            if product_name_cell and str(product_name_cell).strip():
                current_product_name = str(product_name_cell).strip()
                
            if currency_cell and str(currency_cell).strip():
                current_currency = str(currency_cell).strip()
            
            # 檢查是否有關鍵數據
            if transaction_date and quantity is not None and quantity != '':
                # 生成唯一的 transaction_id
                unique_transaction_id = self.generate_unique_transaction_id()
                
                record = {
                    'transaction_id': unique_transaction_id,
                    'customer_id': current_customer_id,
                    'product_id': current_product_id, 
                    'product_name': current_product_name,
                    'transaction_date': transaction_date,
                    'document_type': str(document_type),
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'currency': current_currency,
                    'amount': amount
                }
                
                data.append(record)
        
        logger.info(f"解析完成，共 {len(data)} 筆交易記錄")
        return data
    
    def get_product_is_active(self, product_id: str) -> Optional[bool]:
        """
        從 product_master 表中獲取產品的 is_active 狀態
        
        Args:
            product_id (str): 產品編號
            
        Returns:
            bool: is_active 狀態，如果找不到則返回 None
        """
        try:
            with self.connection.cursor(cursor_factory=RealDictCursor) as cursor:
                query = f"""
                SELECT is_active 
                FROM {self.table_config['product_master']} 
                WHERE product_id = %s
                """
                cursor.execute(query, (product_id,))
                result = cursor.fetchone()
                
                if result:
                    return result['is_active']
                else:
                    logger.warning(f"產品 {product_id} 在 product_master 中未找到")
                    return None
                    
        except Exception as e:
            logger.error(f"查詢產品 {product_id} 的 is_active 狀態失敗: {str(e)}")
            # 回滾當前事務以清除錯誤狀態
            try:
                self.connection.rollback()
            except:
                pass
            return None
    
    def insert_record(self, record: Dict, is_active: Optional[bool]) -> bool:
        """
        插入新記錄
        
        Args:
            record (dict): 交易記錄
            is_active (bool): 產品是否活躍
            
        Returns:
            bool: 是否成功插入（False 表示重複跳過）
        """
        try:
            with self.connection.cursor() as cursor:
                query = f"""
                INSERT INTO {self.table_config['order_transactions']}
                (transaction_id, customer_id, product_id, product_name, 
                 transaction_date, document_type, quantity, unit_price, 
                 currency, amount, is_active, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    record['transaction_id'],
                    record['customer_id'],
                    record['product_id'],
                    record['product_name'],
                    record['transaction_date'],
                    record['document_type'],
                    record['quantity'],
                    record['unit_price'],
                    record['currency'],
                    record['amount'],
                    is_active,
                    datetime.datetime.now()
                ))
                logger.debug(f"插入新記錄 transaction_id: {record['transaction_id']}")
                return True
                
        except psycopg2.errors.UniqueViolation as e:
            # 重複的 transaction_id，跳過這條記錄
            logger.warning(f"跳過重複記錄 transaction_id: {record['transaction_id']}")
            return False
        except Exception as e:
            logger.error(f"插入記錄失敗: {str(e)}")
            raise
    
    def insert_all_records(self, data: List[Dict]) -> int:
        """
        直接插入所有記錄（不檢查重複）
        
        Args:
            data (list): 交易記錄列表
            
        Returns:
            int: 插入記錄數
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        
        inserted_count = 0
        
        try:
            logger.info(f"開始插入 {len(data)} 筆記錄")
            
            for i, record in enumerate(data):
                # 獲取產品的 is_active 狀態
                is_active = self.get_product_is_active(record['product_id'])
                
                # 插入新記錄，如果成功則增加計數
                if self.insert_record(record, is_active):
                    inserted_count += 1
                
                # 每處理一定數量記錄就提交一次
                if (i + 1) % DEFAULT_CONFIG['batch_size'] == 0:
                    self.connection.commit()
                    logger.info(f"已處理 {i + 1} 筆記錄")
            
            # 最終提交
            self.connection.commit()
            logger.info(f"數據插入完成 - 插入: {inserted_count} 筆")
            
            return inserted_count
            
        except Exception as e:
            self.connection.rollback()
            logger.error(f"數據插入失敗: {str(e)}")
            raise
    
    def process_file_with_customer_check(self, file_path: str, delete_month_records: bool = True) -> Tuple[int, int, int, List[str]]:
        """
        處理整個流程：解析文件 -> 過濾有效記錄（只檢查客戶）-> 上傳資料
        
        Args:
            file_path (str): Excel 文件路徑
            delete_month_records (bool): 是否刪除涉及月份記錄，默認為 True
            
        Returns:
            tuple: (刪除記錄數, 插入記錄數, 跳過記錄數, 跳過的客戶列表)
        """
        deleted_count = 0
        
        try:
            # 連接數據庫
            if not self.connect_database():
                raise Exception("無法連接數據庫")
            
            # 1. 解析 Excel 文件
            logger.info("解析 Excel 文件以提取數據...")
            data = self.parse_sales_data(file_path)
            
            # 2. 過濾有效記錄（只檢查客戶，跳過缺失的客戶）
            valid_records, skipped_count, skipped_customers = self.filter_valid_records_by_customer(data)
            
            if skipped_count > 0:
                logger.warning(f"跳過 {skipped_count} 筆記錄，缺失 {len(skipped_customers)} 個客戶")
            
            # 3. 如果沒有有效記錄，直接返回
            if not valid_records:
                logger.warning("沒有有效記錄可以上傳")
                return 0, 0, skipped_count, skipped_customers
            
            # 4. 從有效記錄中提取月份資訊進行刪除
            months_to_delete = self.extract_months_from_data(valid_records)
            logger.info(f"Excel 中包含的月份: {sorted(list(months_to_delete))}")
            
            # 5. 刪除這些月份的所有記錄（如果需要）
            if delete_month_records and months_to_delete:
                logger.info(f"開始刪除 {len(months_to_delete)} 個月份的記錄...")
                deleted_count = self.delete_records_by_months(months_to_delete)
            
            # 6. 插入有效記錄
            inserted_count = self.insert_all_records(valid_records)
            
            return deleted_count, inserted_count, skipped_count, skipped_customers
            
        finally:
            # 關閉數據庫連接
            self.close_connection()

    def process_file(self, file_path: str, delete_month_records: bool = True) -> Tuple[int, int]:
        """
        處理整個流程：解析文件 -> 刪除涉及月份記錄 -> 上傳數據庫
        
        Args:
            file_path (str): Excel 文件路徑
            delete_month_records (bool): 是否刪除涉及月份記錄，默認為 True
            
        Returns:
            tuple: (刪除記錄數, 插入記錄數)
        """
        deleted_count = 0
        
        try:
            # 連接數據庫
            if not self.connect_database():
                raise Exception("無法連接數據庫")
            
            # 1. 先解析 Excel 文件
            logger.info("解析 Excel 文件以提取月份信息...")
            data = self.parse_sales_data(file_path)
            
            # 2. 從數據中提取所有涉及的年月
            months_to_delete = self.extract_months_from_data(data)
            logger.info(f"Excel 中包含的月份: {sorted(list(months_to_delete))}")
            
            # 3. 刪除這些月份的所有記錄（如果需要）
            if delete_month_records and months_to_delete:
                logger.info(f"開始刪除 {len(months_to_delete)} 個月份的記錄...")
                deleted_count = self.delete_records_by_months(months_to_delete)
            
            # 4. 直接插入新記錄
            inserted_count = self.insert_all_records(data)
            
            return deleted_count, inserted_count
            
        finally:
            # 關閉數據庫連接
            self.close_connection()

# API 端點
@router.post("/import/sales/check-customers")
async def check_sales_customers(file: UploadFile = File(...), user_role: str = Form(...)):
    """
    檢查銷貨資料中的客戶是否存在，返回缺失的客戶列表
    """
    try:
        # 檢查權限
        check_editor_permission(user_role)
        
        # 檢查檔案類型
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="只支援 Excel 檔案格式")
        
        # 創建臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        try:
            # 使用 SalesDataUploader 處理檔案
            sales_uploader = SalesDataUploader()
            deleted_count, inserted_count, skipped_count, skipped_customers = sales_uploader.process_file_with_customer_check(temp_file_path)
            
            return JSONResponse(
                status_code=200,
                content={
                    "success": True,
                    "missing_customers": skipped_customers,
                    "message": f"檢查完成，發現 {len(skipped_customers)} 個新客戶需要創建" 
                              if skipped_customers 
                              else f"匯入成功！刪除 {deleted_count} 筆舊記錄，新增 {inserted_count} 筆記錄",
                    "deleted_count": deleted_count,
                    "inserted_count": inserted_count,
                    "filename": file.filename
                }
            )
            
        finally:
            # 清理臨時檔案
            try:
                os.unlink(temp_file_path)
            except:
                pass
                
    except Exception as e:
        logger.error(f"檢查客戶失敗: {str(e)}")
        raise HTTPException(status_code=500, detail=f"檢查失敗: {str(e)}")

@router.post("/import/sales")
async def import_sales_data(file: UploadFile = File(...), user_role: str = Form(...)):
    """
    匯入銷貨資料 API
    """
    try:
        # 檢查權限
        check_editor_permission(user_role)
        
        # 檢查檔案類型
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="只支援 Excel 檔案格式")
        
        # 創建臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        try:
            # 使用 SalesDataUploader 處理檔案
            sales_uploader = SalesDataUploader()
            deleted_count, inserted_count = sales_uploader.process_file(temp_file_path)
            
            return JSONResponse(
                status_code=200,
                content={
                    "success": True,
                    "message": f"匯入成功！刪除 {deleted_count} 筆舊記錄，新增 {inserted_count} 筆交易記錄",
                    "deleted_count": deleted_count,
                    "inserted_count": inserted_count,
                    "filename": file.filename
                }
            )
            
        finally:
            # 清理臨時檔案
            try:
                os.unlink(temp_file_path)
            except:
                pass
                
    except Exception as e:
        logger.error(f"匯入銷貨資料失敗: {str(e)}")
        raise HTTPException(status_code=500, detail=f"匯入失敗: {str(e)}")

@router.post("/customer/create")
async def create_customer(customer_data: dict):
    """
    創建新客戶API
    """
    try:
        uploader = SalesDataUploader()
        if not uploader.connect_database():
            raise Exception("無法連接數據庫")
        
        try:
            # 處理 delivery_schedule 排序 (按 1234567 排序)
            delivery_schedule = customer_data.get('delivery_schedule', '')
            if delivery_schedule:
                # 將字符串轉為列表，排序後再轉回字符串
                schedule_list = delivery_schedule.split(',')
                sorted_schedule = sorted([day for day in schedule_list if day.strip()])
                delivery_schedule = ','.join(sorted_schedule)
            
            # 獲取當前系統時間
            current_time = datetime.datetime.now()
            
            with uploader.connection.cursor() as cursor:
                # 修改 INSERT 語句，加入 created_date 和 updated_date
                query = f"""
                INSERT INTO {uploader.table_config['customer']}
                (customer_id, customer_name, phone_number, address, city, district, delivery_schedule, notes, is_enabled, created_date, updated_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    customer_data['customer_id'],
                    customer_data['customer_name'],
                    customer_data.get('phone_number', ''),
                    customer_data.get('address', ''),
                    customer_data.get('city', ''), 
                    customer_data.get('district', ''),
                    delivery_schedule,  
                    customer_data.get('notes', ''),
                    customer_data.get('is_enabled', 1),
                    current_time,  # created_date
                    current_time   # updated_date
                ))
                uploader.connection.commit()
                
                # 回傳包含時間資訊的回應
                return JSONResponse(
                    status_code=200,
                    content={
                        "success": True,
                        "message": f"客戶 {customer_data['customer_id']} 創建成功",
                        "customer_id": customer_data['customer_id'],
                        "created_date": current_time.isoformat(),
                        "updated_date": current_time.isoformat(),
                        "delivery_schedule": delivery_schedule  # 回傳排序後的配送日程
                    }
                )
                
        finally:
            uploader.close_connection()
            
    except Exception as e:
        logger.error(f"創建客戶失敗: {str(e)}")
        raise HTTPException(status_code=500, detail=f"創建客戶失敗: {str(e)}")

@router.get("/import/test-connection")
async def test_database_connection():
    """
    測試資料庫連接
    """
    try:
        uploader = SalesDataUploader()
        if uploader.connect_database():
            uploader.close_connection()
            return JSONResponse(
                status_code=200,
                content={"success": True, "message": "資料庫連接成功"}
            )
        else:
            return JSONResponse(
                status_code=500,
                content={"success": False, "message": "資料庫連接失敗"}
            )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": f"連接測試失敗: {str(e)}"}
        )
    

@router.post("/import/sales/check-customers-and-products")
async def check_sales_customers(file: UploadFile = File(...), user_role: str = Form(...)):
    """
    檢查銷貨資料中的客戶和產品是否存在，返回缺失的列表
    """
    try:
        # 檢查權限
        check_editor_permission(user_role)
        
        # 檢查檔案類型
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="只支援 Excel 檔案格式")
        
        # 創建臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        try:
            # 使用 SalesDataUploader 處理檔案
            sales_uploader = SalesDataUploader()
            deleted_count, inserted_count, skipped_count, missing_customers, missing_products = sales_uploader.process_file_with_product_check(temp_file_path)
            
            # 構建回應訊息
            if missing_customers or missing_products:
                if inserted_count > 0:
                    message = f"部分匯入成功！刪除 {deleted_count} 筆舊記錄，新增 {inserted_count} 筆交易記錄，跳過 {skipped_count} 筆記錄。發現 {len(missing_customers)} 個新客戶和 {len(missing_products)} 個新產品需要創建"
                else:
                    message = f"檢查完成，發現 {len(missing_customers)} 個新客戶和 {len(missing_products)} 個新產品需要創建"
            else:
                message = f"匯入成功！刪除 {deleted_count} 筆舊記錄，新增 {inserted_count} 筆交易記錄"
            
            return JSONResponse(
                status_code=200,
                content={
                    "success": True,
                    "deleted_count": deleted_count,
                    "inserted_count": inserted_count,
                    "skipped_count": skipped_count,
                    "missing_customers": missing_customers,
                    "missing_products": missing_products,
                    "message": message,
                    "filename": file.filename
                }
            )
            
        finally:
            # 清理臨時檔案
            try:
                os.unlink(temp_file_path)
            except:
                pass
                
    except Exception as e:
        logger.error(f"檢查客戶和產品失敗: {str(e)}")
        raise HTTPException(status_code=500, detail=f"檢查失敗: {str(e)}")

@router.post("/product/create")
async def create_product(product_data: dict):
    """
    創建新產品API
    """
    try:
        uploader = SalesDataUploader()
        if not uploader.connect_database():
            raise Exception("無法連接數據庫")
        
        try:
            # 獲取當前系統時間
            current_time = datetime.datetime.now()
            
            with uploader.connection.cursor() as cursor:
                # 1. 創建產品資料
                product_query = f"""
                INSERT INTO {uploader.table_config['product_master']}
                (product_id, warehouse_id, name_zh, category, subcategory,
                 specification, package_raw, process_type, unit, supplier_id,
                 is_active, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(product_query, (
                    product_data['product_id'],
                    product_data['warehouse_id'],
                    product_data['name_zh'],
                    product_data['category'],
                    product_data['subcategory'],
                    product_data.get('specification', ''),
                    product_data.get('package_raw', ''),
                    product_data.get('process_type', ''),
                    product_data.get('unit', ''),
                    product_data.get('supplier_id', ''),
                    product_data['is_active'],  # 'active' 或 'inactive'
                    current_time,  # created_at
                    current_time   # updated_at
                ))

                # 2. 同時創建庫存資料
                inventory_query = f"""
                INSERT INTO {uploader.table_config['inventory']}
                (product_id, warehouse_id, stock_quantity, total_quantity, unit, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s)
                """
                stock_quantity = product_data.get('stock_quantity', 0)
                total_quantity = product_data.get('total_quantity', stock_quantity)

                cursor.execute(inventory_query, (
                    product_data['product_id'],
                    product_data['warehouse_id'],
                    stock_quantity,
                    total_quantity,
                    product_data.get('unit', ''),
                    current_time
                ))

                uploader.connection.commit()
                
                return JSONResponse(
                    status_code=200,
                    content={
                        "success": True,
                        "message": f"產品 {product_data['product_id']} 及庫存資料創建成功",
                        "product_id": product_data['product_id'],
                        "warehouse_id": product_data['warehouse_id'],
                        "stock_quantity": stock_quantity,
                        "total_quantity": total_quantity,
                        "created_at": current_time.isoformat(),
                        "updated_at": current_time.isoformat()
                    }
                )
                
        finally:
            uploader.close_connection()
            
    except Exception as e:
        logger.error(f"創建產品失敗: {str(e)}")
        raise HTTPException(status_code=500, detail=f"創建產品失敗: {str(e)}")
    
# 庫存上傳
class InventoryDataUploader:
    def __init__(self):
        """初始化庫存數據上傳器"""
        self.connection = None
        
        # 庫存模組專用的表配置 - 根據實際資料庫表結構
        self.table_config = {
            'inventory': 'inventory',  # 實際的庫存表名
            'product_master': 'product_master'
        }
        
    def connect_database(self):
        """連接數據庫"""
        try:
            # 使用新的統一資料庫連接管理
            db_manager = get_db_connection()
            self.connection = db_manager.__enter__()
            self.db_manager = db_manager  # 保存 context manager 以便後續關閉
            logger.info("庫存數據庫連接成功")
            return True
        except Exception as e:
            logger.error(f"庫存數據庫連接失敗: {str(e)}")
            return False
    
    def close_connection(self):
        """關閉數據庫連接"""
        if hasattr(self, 'db_manager') and self.db_manager:
            self.db_manager.__exit__(None, None, None)
            logger.info("庫存數據庫連接已關閉")
    
    def parse_inventory_data(self, file_path: str) -> List[Dict]:
        """
        解析庫存 Excel 文件，提取庫存數據
        
        Args:
            file_path (str): Excel 文件路徑
        
        Returns:
            list: 包含所有庫存記錄的列表
        """
        logger.info(f"開始解析庫存文件: {file_path}")
        
        try:
            # 使用 openpyxl 讀取 Excel 文件
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            logger.info(f"工作表名稱: {worksheet.title}")
            logger.info(f"數據範圍: {worksheet.calculate_dimension()}")
            
            # 獲取最大行數和列數
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            logger.info(f"總行數: {max_row}, 總列數: {max_col}")
            
            # 根據您的 Excel 格式，數據從第8行開始（跳過標題和空行）
            data = []
            
            # 遍歷從第8行開始的所有行
            for row in range(8, max_row + 1):
                # 根據您的 Excel 格式獲取各列的值
                # 根據真實 Excel 格式獲取各列的值
                product_id_cell = worksheet.cell(row=row, column=1).value        # A列: 產品編號
                product_name_cell = worksheet.cell(row=row, column=4).value      # B列: 品名現況（保持原樣）
                category_cell = worksheet.cell(row=row, column=8).value          # C列: 類別（保持原樣）
                warehouse_name_cell = worksheet.cell(row=row, column=10).value   # J列: 倉庫名稱
                total_quantity_cell = worksheet.cell(row=row, column=13).value   # M列: 數量
                borrowed_out_cell = worksheet.cell(row=row, column=15).value     # O列: 借出數量
                borrowed_in_cell = worksheet.cell(row=row, column=18).value      # R列: 借入數量
                stock_quantity_cell = worksheet.cell(row=row, column=20).value   # T列: 實際在庫量
                unit_cell = worksheet.cell(row=row, column=23).value            # W列: 單位
                
                # 檢查是否有關鍵數據 (產品編號必須存在)
                if product_id_cell and str(product_id_cell).strip():
                    product_id = str(product_id_cell).strip()
                    logger.debug(f"處理產品: {product_id}")
                    
                    
                    # 處理倉庫名稱
                    warehouse_id = ""
                    if warehouse_name_cell and str(warehouse_name_cell).strip():
                        warehouse_id = str(warehouse_name_cell).strip()
                    
                    # 處理數量欄位
                    total_quantity = 0
                    borrowed_out = 0  
                    borrowed_in = 0
                    stock_quantity = 0
                    
                    try:
                        if total_quantity_cell is not None and str(total_quantity_cell).strip():
                            # 移除可能的文字並轉換為數字
                            total_quantity_str = str(total_quantity_cell).replace(',', '')
                            if total_quantity_str != '---' and total_quantity_str != '':
                                total_quantity = float(total_quantity_str)
                                
                        if borrowed_out_cell is not None and str(borrowed_out_cell).strip():
                            borrowed_out_str = str(borrowed_out_cell).replace(',', '')
                            if borrowed_out_str != '---' and borrowed_out_str != '':
                                borrowed_out = float(borrowed_out_str)
                                
                        if borrowed_in_cell is not None and str(borrowed_in_cell).strip():
                            borrowed_in_str = str(borrowed_in_cell).replace(',', '')
                            if borrowed_in_str != '---' and borrowed_in_str != '':
                                borrowed_in = float(borrowed_in_str)
                                
                        if stock_quantity_cell is not None and str(stock_quantity_cell).strip():
                            stock_quantity_str = str(stock_quantity_cell).replace(',', '')
                            if stock_quantity_str != '---' and stock_quantity_str != '':
                                stock_quantity = float(stock_quantity_str)
                                
                    except (ValueError, TypeError) as e:
                        logger.warning(f"數量轉換錯誤，行 {row}: {str(e)}")
                        continue

                    # 處理單位
                    unit = ""
                    if unit_cell and str(unit_cell).strip():
                        unit = str(unit_cell).strip()
                    
                    # 處理品名規格
                    product_name_zh = ""
                    if product_name_cell and str(product_name_cell).strip():
                        product_name_zh = str(product_name_cell).strip()
                    
                    # 處理類別
                    category = ""
                    if category_cell and str(category_cell).strip():
                        category = str(category_cell).strip()
                    
                    # 創建記錄
                    record = {
                        'product_id': product_id,
                        'warehouse_id': warehouse_id,
                        'total_quantity': total_quantity,
                        'borrowed_out': borrowed_out,
                        'borrowed_in': borrowed_in,
                        'stock_quantity': stock_quantity,
                        'unit': unit,
                        'product_name_zh': product_name_zh,
                        'category': category
                    }

                    data.append(record)
                    logger.debug(f"新增庫存記錄: 產品ID={product_id}, 倉庫={warehouse_id}, 庫存量={stock_quantity}")
                else:
                    logger.debug(f"第 {row} 行跳過 - 無產品編號")
            
            logger.info(f"庫存解析完成，共 {len(data)} 筆有效記錄")
            return data
            
        except Exception as e:
            logger.error(f"解析庫存文件失敗: {str(e)}")
            raise
    
    def check_product_exists(self, product_id: str) -> bool:
        """
        檢查產品是否已存在於 product_master 表中
        
        Args:
            product_id (str): 產品ID
            
        Returns:
            bool: 如果已存在返回 True，否則返回 False
        """
        try:
            with self.connection.cursor() as cursor:
                query = f"""
                SELECT EXISTS(
                    SELECT 1 FROM {self.table_config['product_master']}
                    WHERE product_id = %s
                )
                """
                cursor.execute(query, (product_id,))
                return cursor.fetchone()[0]
                
        except Exception as e:
            logger.error(f"檢查產品存在性失敗: {str(e)}")
            try:
                self.connection.rollback()
            except:
                pass
            return False

    def get_missing_products(self, data: List[Dict]) -> List[Dict]:
        """
        檢查數據中哪些產品不存在於 product_master 表中
        
        Args:
            data (list): 庫存記錄列表
            
        Returns:
            list: 不存在的產品資訊列表，包含從Excel提取的欄位
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        def safe_str(value):
            return str(value).strip() if value is not None else ""

        
        # 收集唯一的產品資料，包含Excel中的相關欄位
        unique_products = {}
        for record in data:
            product_id = record.get('product_id')
            if product_id and product_id.strip():
                product_id = product_id.strip()
                if product_id not in unique_products:
                    # 從Excel記錄中提取產品資訊
                    unique_products[product_id] = {
                        'product_id': product_id,
                        'name_zh': safe_str(record.get('product_name_zh')),
                        'category': safe_str(record.get('category')),
                        'unit': safe_str(record.get('unit')),
                        'warehouse_id': safe_str(record.get('warehouse_id'))
                    }
        
        missing_products = []
        
        for product_id, product_info in unique_products.items():
            if not self.check_product_exists(product_id):
                missing_products.append(product_info)
                logger.info(f"發現新產品: {product_id} - {product_info.get('name_zh', '')}")
        
        return missing_products

    def filter_valid_inventory_records(self, data: List[Dict]) -> Tuple[List[Dict], int, List[str]]:
        """
        過濾出有效的庫存記錄（產品存在的記錄）
        
        Args:
            data (list): 原始庫存記錄列表
            
        Returns:
            tuple: (有效記錄列表, 跳過記錄數, 跳過的產品ID列表)
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        
        valid_records = []
        skipped_products = set()
        skipped_count = 0
        
        for record in data:
            product_id = record.get('product_id', '').strip()
            
            # 檢查產品是否存在  
            product_exists = self.check_product_exists(product_id) if product_id else False
            
            if product_exists:
                # 產品存在，記錄有效
                valid_records.append(record)
            else:
                # 記錄跳過的產品
                skipped_count += 1
                if product_id:
                    skipped_products.add(product_id)
                    
                logger.info(f"跳過庫存記錄: product_id={product_id}, product_exists={product_exists}")
        
        logger.info(f"庫存記錄過濾完成: 有效={len(valid_records)}, 跳過={skipped_count}")
        return valid_records, skipped_count, list(skipped_products)

    def insert_inventory_records(self, data: List[Dict]) -> int:
        """
        插入庫存記錄
        """
        if not self.connection:
            raise Exception("數據庫未連接")
        
        inserted_count = 0
        
        try:
            logger.info(f"開始插入 {len(data)} 筆庫存記錄")
            
            for i, record in enumerate(data):
                # 直接插入新記錄
                self.insert_inventory_record(record)  # ← 改成調用單獨的插入方法
                inserted_count += 1
                
                # 每處理一定數量記錄就提交一次
                if (i + 1) % DEFAULT_CONFIG['batch_size'] == 0:
                    self.connection.commit()
                    logger.info(f"已處理 {i + 1} 筆庫存記錄")
            
            # 最終提交
            self.connection.commit()
            logger.info(f"庫存數據插入完成 - 插入: {inserted_count} 筆")
            
            return inserted_count
            
        except Exception as e:
            self.connection.rollback()
            logger.error(f"庫存數據插入失敗: {str(e)}")
            raise
    def insert_inventory_record(self, record: Dict):
        """
        插入單筆庫存記錄
        
        Args:
            record (dict): 庫存記錄
        """
        try:
            with self.connection.cursor() as cursor:
                query = f"""
                INSERT INTO {self.table_config['inventory']}
                (product_id, warehouse_id, total_quantity, borrowed_out, borrowed_in, 
                stock_quantity, unit, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    record['product_id'],
                    record['warehouse_id'],
                    record['total_quantity'],
                    record['borrowed_out'],
                    record['borrowed_in'],
                    record['stock_quantity'],
                    record['unit'],
                    datetime.datetime.now(),  # created_at
                    datetime.datetime.now()
                ))
                logger.debug(f"插入庫存記錄 product_id: {record['product_id']}")
                
        except Exception as e:
            logger.error(f"插入庫存記錄失敗: {str(e)}")
            raise
    def process_file(self, file_path: str, replace_existing: bool = True) -> Tuple[int, int]:
            """
            處理庫存文件：解析 -> 刪除舊記錄 -> 插入新記錄
            
            Args:
                file_path (str): Excel 文件路徑
                replace_existing (bool): 是否替換現有記錄，默認為 True
                
            Returns:
                tuple: (刪除記錄數, 插入記錄數)
            """
            deleted_count = 0
            
            try:
                # 連接數據庫
                if not self.connect_database():
                    raise Exception("無法連接數據庫")
                
                # 1. 解析 Excel 文件
                logger.info("解析庫存 Excel 文件...")
                data = self.parse_inventory_data(file_path)
                
                if not data:
                    logger.warning("沒有解析到任何有效的庫存記錄")
                    return 0, 0
                
                if replace_existing and data:
                    # 2. 提取所有產品ID和倉庫ID
                    product_ids = list(set([record['product_id'] for record in data if record['product_id']]))
                    warehouse_ids = list(set([record['warehouse_id'] for record in data if record['warehouse_id']]))
                    logger.info(f"準備更新 {len(product_ids)} 個產品在 {len(warehouse_ids)} 個倉庫的庫存")
                    
                    # 3. 刪除現有記錄
                    deleted_count = self.delete_existing_inventory(product_ids, warehouse_ids)
                
                # 4. 插入新記錄
                inserted_count = self.insert_inventory_records(data)
                
                return deleted_count, inserted_count
                
            finally:
                # 關閉數據庫連接
                self.close_connection()
    def delete_existing_inventory(self, product_ids: List[str], warehouse_ids: List[str] = None) -> int:
            """
            刪除現有的庫存記錄
            
            Args:
                product_ids (list): 要刪除的產品ID列表
                warehouse_ids (list): 要刪除的倉庫ID列表（可選）
                
            Returns:
                int: 刪除的記錄數
            """
            if not product_ids:
                return 0
                
            deleted_count = 0
            
            try:
                with self.connection.cursor() as cursor:
                    if warehouse_ids:
                        # 如果指定倉庫，只刪除特定產品在特定倉庫的記錄
                        product_placeholders = ','.join(['%s'] * len(product_ids))
                        warehouse_placeholders = ','.join(['%s'] * len(warehouse_ids))
                        query = f"""
                        DELETE FROM {self.table_config['inventory']}
                        WHERE product_id IN ({product_placeholders})
                        AND warehouse_id IN ({warehouse_placeholders})
                        """
                        cursor.execute(query, product_ids + warehouse_ids)
                    else:
                        # 刪除所有相關產品的庫存記錄
                        placeholders = ','.join(['%s'] * len(product_ids))
                        query = f"""
                        DELETE FROM {self.table_config['inventory']}
                        WHERE product_id IN ({placeholders})
                        """
                        cursor.execute(query, product_ids)
                    
                    deleted_count = cursor.rowcount
                    logger.info(f"刪除了 {deleted_count} 筆庫存記錄")
                
                self.connection.commit()
                return deleted_count
                    
            except Exception as e:
                self.connection.rollback()
                logger.error(f"刪除庫存記錄失敗: {str(e)}")
                raise
    def process_file_with_product_check(self, file_path: str, replace_existing: bool = True) -> Tuple[int, int, int, List[Dict]]:
        """
        處理庫存文件：解析文件 -> 上傳存在的記錄 -> 返回缺失產品列表
        
        Args:
            file_path (str): Excel 文件路徑
            replace_existing (bool): 是否替換現有記錄，默認為 True
            
        Returns:
            tuple: (deleted_count, inserted_count, skipped_count, missing_products_details)
        """
        deleted_count = 0
        
        try:
            # 連接數據庫
            if not self.connect_database():
                raise Exception("無法連接數據庫")
            
            # 1. 解析 Excel 文件
            logger.info("解析庫存 Excel 文件...")
            data = self.parse_inventory_data(file_path)
            
            if not data:
                logger.warning("沒有解析到任何有效的庫存記錄")
                return 0, 0, 0, []
            
            # 2. 過濾有效記錄（混合模式：上傳存在的，返回缺失的）
            valid_records, skipped_count, skipped_products = self.filter_valid_inventory_records(data)
            missing_products = self.get_missing_products(data)

            
            if skipped_count > 0:
                logger.info(f"將上傳 {len(valid_records)} 筆有效庫存記錄，跳過 {skipped_count} 筆記錄（缺失 {len(skipped_products)} 個產品）")
            
            # 3. 上傳有效記錄（即使有跳過的記錄也要上傳）
            if valid_records:
                if replace_existing:
                    # 從有效記錄中提取所有產品ID和倉庫ID
                    product_ids = list(set([record['product_id'] for record in valid_records if record['product_id']]))
                    warehouse_ids = list(set([record['warehouse_id'] for record in valid_records if record['warehouse_id']]))
                    logger.info(f"準備更新 {len(product_ids)} 個產品在 {len(warehouse_ids)} 個倉庫的庫存")
                    
                    # 刪除現有記錄
                    deleted_count = self.delete_existing_inventory(product_ids, warehouse_ids)
                
                # 插入有效記錄
                inserted_count = self.insert_inventory_records(valid_records)
            else:
                inserted_count = 0
            
            # 4. 返回結果：包含上傳統計和缺失產品
            return deleted_count, inserted_count, skipped_count, missing_products
            
        finally:
            # 關閉數據庫連接
            self.close_connection()



# 庫存API端點
@router.post("/import/inventory")
async def import_inventory_data(file: UploadFile = File(...), user_role: str = Form(...)):
    """
    匯入庫存資料 API
    """
    try:
        # 檢查權限
        check_editor_permission(user_role)
        
        # 檢查檔案類型
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="只支援 Excel 檔案格式")
        
        # 創建臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        try:
            # 使用 InventoryDataUploader 處理檔案
            inventory_uploader = InventoryDataUploader()
            deleted_count, inserted_count = inventory_uploader.process_file(temp_file_path)
            
            return JSONResponse(
                status_code=200,
                content={
                    "success": True,
                    "message": f"庫存匯入成功！刪除 {deleted_count} 筆舊記錄，新增 {inserted_count} 筆庫存記錄",
                    "deleted_count": deleted_count,
                    "inserted_count": inserted_count,
                    "filename": file.filename
                }
            )
            
        finally:
            # 清理臨時檔案
            try:
                os.unlink(temp_file_path)
            except:
                pass
                
    except Exception as e:
        logger.error(f"匯入庫存資料失敗: {str(e)}")
        raise HTTPException(status_code=500, detail=f"匯入失敗: {str(e)}")



# 在檔案最後新增這個 API 端點
@router.post("/import/inventory/check-products")
async def check_inventory_products(file: UploadFile = File(...), user_role: str = Form(...)):
    """
    檢查庫存資料中的產品是否存在，返回缺失的產品列表
    """
    try:
        # 檢查權限
        check_editor_permission(user_role)
        
        # 檢查檔案類型
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="只支援 Excel 檔案格式")
        
        # 創建臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        try:
            # 使用 InventoryDataUploader 處理檔案
            inventory_uploader = InventoryDataUploader()
            deleted_count, inserted_count, skipped_count, missing_products = inventory_uploader.process_file_with_product_check(temp_file_path)
            
            # 構建回應訊息
            if missing_products:
                if inserted_count > 0:
                    message = f"部分匯入成功！刪除 {deleted_count} 筆舊記錄，新增 {inserted_count} 筆庫存記錄，跳過 {skipped_count} 筆記錄。發現 {len(missing_products)} 個新產品需要創建"
                else:
                    message = f"檢查完成，發現 {len(missing_products)} 個新產品需要創建"
            else:
                message = f"庫存匯入成功！刪除 {deleted_count} 筆舊記錄，新增 {inserted_count} 筆庫存記錄"
            
            return JSONResponse(
                status_code=200,
                content={
                    "success": True,
                    "deleted_count": deleted_count,
                    "inserted_count": inserted_count,
                    "skipped_count": skipped_count,
                    "missing_products": missing_products,
                    "message": message,
                    "filename": file.filename
                }
            )
            
        finally:
            # 清理臨時檔案
            try:
                os.unlink(temp_file_path)
            except:
                pass
                
    except Exception as e:
        logger.error(f"檢查庫存產品失敗: {str(e)}")
        raise HTTPException(status_code=500, detail=f"檢查失敗: {str(e)}")
